"""
=====================================================================
 SCRIPT 3: MÉTRICAS RAGAS
=====================================================================

Calcula las métricas RAGAS sobre los resultados del pipeline RAG:
    Métricas de RECUPERACIÓN:
        - Context Precision
        - Context Recall
    Métricas de GENERACIÓN:
        - Faithfulness
        - Answer Relevancy
        - Answer Correctness

EXCLUSIONES:
    - D_jerarquico: solo se evalúan métricas de generación
      (las de recuperación se inflan artificialmente por el tamaño
       de los padres, ver memoria)
    - Negativas: se evalúan aparte porque el ground truth es distinto
      (la respuesta correcta es "no sé" / declinar la respuesta)

CONFIGURACIÓN:
    - LLM juez: Gemini 2.5 Flash vía API
    - Embeddings: paraphrase-multilingual-mpnet-base-v2 (neutral)
    - Concurrencia: 8 peticiones en paralelo (configurable)

Input:
    - resultados_retrieval_generacion.json (output script 1)

Output:
    - metricas_ragas.json (resultados completos por ejecución)
    - metricas_ragas_resumen.xlsx (tablas agregadas)
"""

import os
import json
import time
import warnings
from pathlib import Path

warnings.filterwarnings('ignore')
os.environ['HF_HUB_DISABLE_PROGRESS_BARS'] = '1'
os.environ['TOKENIZERS_PARALLELISM'] = 'false'

import pandas as pd
from datasets import Dataset
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_huggingface import HuggingFaceEmbeddings

from ragas import evaluate
from ragas.llms import LangchainLLMWrapper
from ragas.embeddings import LangchainEmbeddingsWrapper
from ragas.metrics import (
    context_precision,
    context_recall,
    faithfulness,
    answer_relevancy,
    answer_correctness,
)
from ragas.run_config import RunConfig

# ═══════════════════════════════════════════════════════════════════
# CONFIGURACIÓN
# ═══════════════════════════════════════════════════════════════════

# Archivos
ARCHIVO_RESULTADOS_PIPELINE = 'resultados_retrieval_generacion.json'
ARCHIVO_SALIDA_JSON = 'metricas_ragas.json'
ARCHIVO_SALIDA_EXCEL = 'metricas_ragas_resumen.xlsx'

# Modo test: si True, solo procesa unas pocas ejecuciones para verificar
# que todo funciona antes de lanzar el experimento completo.
MODO_TEST = False
N_TEST = 5

# API
GEMINI_API_KEY = "AIzaSyAF4ZZ2jjC1nzvxWOgxXh8IP1YJlF6gzCc"

# Modelos
MODELO_LLM_JUEZ = 'gemini-2.5-flash'
MODELO_EMBEDDINGS_JUEZ = 'sentence-transformers/paraphrase-multilingual-mpnet-base-v2'

# Concurrencia y reintentos
MAX_WORKERS = 8
MAX_RETRIES = 5
TIMEOUT_SEGUNDOS = 180

# Lotes para checkpoint incremental
TAMANO_LOTE = 50  # nº de ejecuciones a evaluar antes de cada guardado

# Filtros
CHUNKINGS_RECUPERACION = ['A_fixed', 'B_markdown', 'C_semantica']  # D excluido
CHUNKINGS_GENERACION = ['A_fixed', 'B_markdown', 'C_semantica', 'D_jerarquico']
TIPOS_EVALUADOS = ['Directa', 'Caso', 'Compleja']  # Negativas se evalúan aparte


# ═══════════════════════════════════════════════════════════════════
# INICIALIZACIÓN DE MODELOS RAGAS
# ═══════════════════════════════════════════════════════════════════

print("═" * 70)
print(" CÁLCULO DE MÉTRICAS RAGAS")
print("═" * 70)

print(" - Inicializando LLM juez (Gemini 2.5 Flash)...")
llm_base = ChatGoogleGenerativeAI(
    model=MODELO_LLM_JUEZ,
    google_api_key=GEMINI_API_KEY,
    temperature=0.0,
    max_retries=MAX_RETRIES,
    timeout=TIMEOUT_SEGUNDOS,
)
llm_juez = LangchainLLMWrapper(llm_base)

print(f" - Inicializando embeddings ({MODELO_EMBEDDINGS_JUEZ})...")
emb_base = HuggingFaceEmbeddings(model_name=MODELO_EMBEDDINGS_JUEZ)
embeddings_juez = LangchainEmbeddingsWrapper(emb_base)

run_config = RunConfig(
    timeout=TIMEOUT_SEGUNDOS,
    max_retries=MAX_RETRIES,
    max_workers=MAX_WORKERS,
)
print(f" - Concurrencia: {MAX_WORKERS} workers")
print()


# ═══════════════════════════════════════════════════════════════════
# CARGA DE DATOS
# ═══════════════════════════════════════════════════════════════════

with open(ARCHIVO_RESULTADOS_PIPELINE, 'r', encoding='utf-8') as f:
    resultados_pipeline = json.load(f)
print(f" - Resultados pipeline cargados: {len(resultados_pipeline)} ejecuciones")


# ═══════════════════════════════════════════════════════════════════
# CHECKPOINTING
# ═══════════════════════════════════════════════════════════════════

def cargar_checkpoint():
    if not os.path.exists(ARCHIVO_SALIDA_JSON):
        return []
    with open(ARCHIVO_SALIDA_JSON, 'r', encoding='utf-8') as f:
        return json.load(f)


def guardar_checkpoint(metricas):
    with open(ARCHIVO_SALIDA_JSON, 'w', encoding='utf-8') as f:
        json.dump(metricas, f, ensure_ascii=False, indent=2)


def clave_ejecucion(r):
    return f"{r['pregunta_idx']}__{r['chunking']}__{r['retrieval']}"


# ═══════════════════════════════════════════════════════════════════
# CONSTRUCCIÓN DEL DATASET RAGAS
# ═══════════════════════════════════════════════════════════════════

def construir_dataset_ragas(ejecuciones, incluir_recuperacion):
    """
    Convierte la lista de ejecuciones del pipeline al formato RAGAS.
    Si incluir_recuperacion=False, no incluye contextos (usado solo
    para D_jerarquico, que no evaluamos en métricas de recuperación).
    """
    datos = {
        'user_input': [],
        'retrieved_contexts': [],
        'response': [],
        'reference': [],
    }
    for r in ejecuciones:
        datos['user_input'].append(r['pregunta'])
        datos['retrieved_contexts'].append(
            [ch['text'] for ch in r['chunks_recuperados']]
        )
        datos['response'].append(r['respuesta_generada'])
        datos['reference'].append(r['respuesta_ideal'])
    return Dataset.from_dict(datos)


# ═══════════════════════════════════════════════════════════════════
# EVALUACIÓN POR LOTES
# ═══════════════════════════════════════════════════════════════════

def evaluar_lote(ejecuciones, metricas_a_calcular):
    """Evalúa un lote de ejecuciones con las métricas dadas."""
    dataset = construir_dataset_ragas(ejecuciones, incluir_recuperacion=True)
    resultado = evaluate(
        dataset=dataset,
        metrics=metricas_a_calcular,
        llm=llm_juez,
        embeddings=embeddings_juez,
        run_config=run_config,
        show_progress=True,
        raise_exceptions=False,
    )
    return resultado.to_pandas()


def integrar_resultados(ejecuciones_lote, df_resultado, evaluadas_recup):
    """
    Integra los resultados de RAGAS con la metadata de cada ejecución.
    Mapea por orden (RAGAS conserva el orden del dataset).
    """
    salidas = []
    for i, r in enumerate(ejecuciones_lote):
        salida = {
            'pregunta_idx': r['pregunta_idx'],
            'tipo': r['tipo'],
            'source': r['source'],
            'chunking': r['chunking'],
            'retrieval': r['retrieval'],
        }

        # Métricas presentes en el DataFrame (algunas faltan si fallaron)
        for col in df_resultado.columns:
            if col in {'user_input', 'retrieved_contexts', 'response', 'reference'}:
                continue
            valor = df_resultado.iloc[i][col]
            if pd.isna(valor):
                salida[col] = None
            else:
                salida[col] = float(valor)

        # Si esta ejecución es D_jerarquico, sus métricas de recuperación
        # se marcan como None explícitamente (las pasamos por el evaluador
        # solo para mantener un orden coherente, pero las descartamos)
        if not evaluadas_recup:
            salida['context_precision'] = None
            salida['context_recall'] = None

        salidas.append(salida)
    return salidas


# ═══════════════════════════════════════════════════════════════════
# LOOP PRINCIPAL
# ═══════════════════════════════════════════════════════════════════

def main():
    # Filtrar ejecuciones a evaluar (excluir Negativas; D se evalúa aparte)
    ejecuciones_validas = [
        r for r in resultados_pipeline
        if r['tipo'] in TIPOS_EVALUADOS
    ]
    n_negativas = len(resultados_pipeline) - len(ejecuciones_validas)

    # En modo test, solo unas pocas
    if MODO_TEST:
        ejecuciones_validas = ejecuciones_validas[:N_TEST]

    # Separar por chunking: las de A/B/C van con todas las métricas;
    # las de D solo con las de generación
    eje_recup_y_gen = [r for r in ejecuciones_validas
                       if r['chunking'] in CHUNKINGS_RECUPERACION]
    eje_solo_gen = [r for r in ejecuciones_validas
                    if r['chunking'] not in CHUNKINGS_RECUPERACION]

    print(f" - Total resultados pipeline:           {len(resultados_pipeline)}")
    print(f" - Excluidas por tipo (Negativa):       {n_negativas}")
    print(f" - A evaluar:                           {len(ejecuciones_validas)}")
    print(f"     · Con métricas recup. + generac.:  {len(eje_recup_y_gen)}")
    print(f"     · Solo generación (D_jerarquico):  {len(eje_solo_gen)}")
    if MODO_TEST:
        print(f"\n   *** MODO TEST: solo {N_TEST} ejecuciones ***\n")

    # Cargar checkpoint
    metricas_previas = cargar_checkpoint()
    hechas = {f"{m['pregunta_idx']}__{m['chunking']}__{m['retrieval']}"
              for m in metricas_previas}
    print(f" - Ya evaluadas (checkpoint): {len(hechas)}")
    print()

    metricas_globales = list(metricas_previas)
    t_inicio = time.time()

    # ── BLOQUE 1: A/B/C con todas las métricas ──────────────────────
    metricas_completas = [
        context_precision,
        context_recall,
        faithfulness,
        answer_relevancy,
        answer_correctness,
    ]
    pendientes_recup = [r for r in eje_recup_y_gen if clave_ejecucion(r) not in hechas]
    print(f" • Procesando {len(pendientes_recup)} ejecuciones (A/B/C)...")

    for i in range(0, len(pendientes_recup), TAMANO_LOTE):
        lote = pendientes_recup[i:i + TAMANO_LOTE]
        n_lote = len(lote)
        print(f"\n   Lote {i//TAMANO_LOTE + 1}/"
              f"{(len(pendientes_recup) - 1)//TAMANO_LOTE + 1} "
              f"({n_lote} ejecuciones)")
        try:
            df_resultado = evaluar_lote(lote, metricas_completas)
            integradas = integrar_resultados(lote, df_resultado, evaluadas_recup=True)
            metricas_globales.extend(integradas)
            guardar_checkpoint(metricas_globales)
            print(f"     ✓ Lote completado y guardado ({len(metricas_globales)} totales)")
        except Exception as e:
            print(f"     [ERROR] {type(e).__name__}: {e}")
            guardar_checkpoint(metricas_globales)
            continue

    # ── BLOQUE 2: D_jerarquico solo con métricas de generación ──────
    metricas_solo_gen = [
        faithfulness,
        answer_relevancy,
        answer_correctness,
    ]
    pendientes_d = [r for r in eje_solo_gen if clave_ejecucion(r) not in hechas]
    if pendientes_d:
        print(f"\n • Procesando {len(pendientes_d)} ejecuciones (D_jerarquico)...")
        for i in range(0, len(pendientes_d), TAMANO_LOTE):
            lote = pendientes_d[i:i + TAMANO_LOTE]
            n_lote = len(lote)
            print(f"\n   Lote D {i//TAMANO_LOTE + 1}/"
                  f"{(len(pendientes_d) - 1)//TAMANO_LOTE + 1} "
                  f"({n_lote} ejecuciones)")
            try:
                df_resultado = evaluar_lote(lote, metricas_solo_gen)
                integradas = integrar_resultados(lote, df_resultado, evaluadas_recup=False)
                metricas_globales.extend(integradas)
                guardar_checkpoint(metricas_globales)
                print(f"     ✓ Lote completado y guardado ({len(metricas_globales)} totales)")
            except Exception as e:
                print(f"     [ERROR] {type(e).__name__}: {e}")
                guardar_checkpoint(metricas_globales)
                continue

    # Guardado final
    guardar_checkpoint(metricas_globales)

    t_total = (time.time() - t_inicio) / 60
    print(f"\n - Tiempo transcurrido: {t_total:.1f} min")
    print(f" - Métricas en archivo: {len(metricas_globales)}")
    print(f" - Archivo: {ARCHIVO_SALIDA_JSON}")

    # ── AGREGACIÓN Y EXCEL ──────────────────────────────────────────
    if metricas_globales:
        generar_excel(metricas_globales)


# ═══════════════════════════════════════════════════════════════════
# AGREGACIÓN + EXCEL
# ═══════════════════════════════════════════════════════════════════

def generar_excel(metricas):
    df = pd.DataFrame(metricas)

    cols_metricas = [c for c in [
        'context_precision', 'context_recall',
        'faithfulness', 'answer_relevancy', 'answer_correctness',
    ] if c in df.columns]

    def agregar(df_in, agrupar_por):
        return (df_in
                .groupby(agrupar_por)[cols_metricas]
                .mean()
                .round(4)
                .reset_index())

    df_global = agregar(df, ['chunking', 'retrieval'])
    df_por_tipo = agregar(df, ['chunking', 'retrieval', 'tipo'])
    df_por_chunking = agregar(df, ['chunking'])
    df_por_retrieval = agregar(df, ['retrieval'])

    print()
    print("═" * 70)
    print(" RESUMEN GLOBAL (chunking × retrieval)")
    print("═" * 70)
    print(df_global.to_string(index=False))

    # Excel
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    thin = Side(style='thin', color='BBBBBB')
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    color_titulo = '1F3864'
    color_header = 'CCCCCC'

    def estilo(cell, bold=False, bg=None, color='000000'):
        cell.font = Font(name='Arial', bold=bold, color=color, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = borde
        if bg:
            cell.fill = PatternFill('solid', start_color=bg)

    def añadir_hoja(nombre, df_h, titulo):
        ws = wb.create_sheet(nombre)
        n_cols = len(df_h.columns)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        c = ws.cell(row=1, column=1, value=titulo)
        estilo(c, bold=True, bg=color_titulo, color='FFFFFF')
        ws.row_dimensions[1].height = 24
        for col, cab in enumerate(df_h.columns, 1):
            estilo(ws.cell(row=2, column=col, value=str(cab)),
                   bold=True, bg=color_header)
            ws.column_dimensions[get_column_letter(col)].width = 16
        for fila, (_, row) in enumerate(df_h.iterrows(), 3):
            for col, val in enumerate(row, 1):
                if isinstance(val, float):
                    val = round(val, 4) if pd.notna(val) else '-'
                estilo(ws.cell(row=fila, column=col, value=val))

    añadir_hoja('Global', df_global, 'MÉTRICAS RAGAS — VISTA GLOBAL')
    añadir_hoja('Por Tipo', df_por_tipo, 'MÉTRICAS RAGAS — DESGLOSE POR TIPO')
    añadir_hoja('Por Chunking', df_por_chunking, 'MÉTRICAS RAGAS — POR CHUNKING')
    añadir_hoja('Por Retrieval', df_por_retrieval, 'MÉTRICAS RAGAS — POR RETRIEVAL')

    wb.save(ARCHIVO_SALIDA_EXCEL)
    print(f"\n - Excel guardado: {ARCHIVO_SALIDA_EXCEL}")


if __name__ == '__main__':
    main()
