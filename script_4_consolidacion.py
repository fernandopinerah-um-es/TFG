"""
=====================================================================
 SCRIPT 4: CONSOLIDACIÓN Y VISUALIZACIONES FINALES DEL TFG
=====================================================================

Genera todos los artefactos finales para la memoria del TFG con
DOS RANKINGS SEPARADOS por decisión metodológica:

  - Ranking PRINCIPAL (A/B/C × 7 retrievals = 21 combinaciones)
    Score compuesto con TODAS las métricas IR + RAGAS.
    Responde: ¿qué combinación end-to-end es la mejor?

  - Ranking de GENERACIÓN (A/B/C/D × 7 retrievals = 28 combinaciones)
    Score compuesto solo con métricas RAGAS de generación
    (Faithfulness, Answer Relevancy, Answer Correctness).
    Responde: ¿qué chunking produce las mejores respuestas?
              ¿Aporta D_jerarquico algo en calidad de respuesta?

Manejo de nulos:
    - Los `null` literales del JSON se cargan como NaN de pandas.
    - .mean() de pandas IGNORA NaN automáticamente: registros con
      NaN en una métrica NO se descartan, solo no aportan a esa
      métrica.
    - El score compuesto usa fillna(0) sobre la tabla agregada.
      Si una celda agregada es NaN (todas las ejecuciones eran NaN),
      cuenta como 0 → penaliza la combinación. Por eso el ranking
      principal excluye D_jerarquico (sus context_precision/recall
      son NaN por diseño y serían penalizadas injustamente).

Input:
    - resultados_retrieval_generacion.json (script 1)
    - metricas_ir.json                     (script 2)
    - metricas_ragas.json                  (script 3)
    - metricas_negativas.json              (script 3b)

Output:
    - resultados_finales.xlsx
    - graficas/ (carpeta con todos los PNG)
    - resumen_ejecutivo.md
    - diagnostico_calidad_ragas.json
"""

import os
import json
import warnings
from pathlib import Path
from collections import defaultdict

warnings.filterwarnings('ignore')

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ═══════════════════════════════════════════════════════════════════
# CONFIGURACIÓN
# ═══════════════════════════════════════════════════════════════════

# Inputs
ARCHIVO_PIPELINE = 'resultados_retrieval_generacion.json'
ARCHIVO_IR = 'metricas_ir.json'
ARCHIVO_RAGAS = 'metricas_ragas.json'
ARCHIVO_NEGATIVAS = 'metricas_negativas.json'

# Outputs
DIR_GRAFICAS = Path('graficas')
ARCHIVO_EXCEL = 'resultados_finales.xlsx'
ARCHIVO_RESUMEN = 'resumen_ejecutivo.md'

# Estilo de gráficos (calidad publicable)
plt.rcParams.update({
    'figure.dpi': 150,
    'savefig.dpi': 300,
    'savefig.bbox': 'tight',
    'savefig.facecolor': 'white',
    'font.family': 'sans-serif',
    'font.size': 11,
    'axes.titlesize': 13,
    'axes.labelsize': 11,
    'axes.spines.top': False,
    'axes.spines.right': False,
    'axes.grid': True,
    'axes.axisbelow': True,
    'grid.alpha': 0.3,
})

# Paleta consistente
PALETA_CHUNKING = {
    'A_fixed': '#1f77b4',
    'B_markdown': '#2ca02c',
    'C_semantica': '#ff7f0e',
    'D_jerarquico': '#d62728',
}
PALETA_RETRIEVAL = {
    'bm25': '#4C72B0',
    'semantica_generalista': '#55A868',
    'semantica_medico': '#C44E52',
    'semantica_experto': '#8172B2',
    'hibrida_generalista': '#CCB974',
    'hibrida_medico': '#64B5CD',
    'hibrida_experto': '#937860',
}

# ─── PESOS DEL RANKING PRINCIPAL (A/B/C, end-to-end) ───
# Suma = 1.00. Equilibra IR (0.30) + retrieval RAGAS (0.30) + generación (0.40).
PESOS_PRINCIPAL = {
    'mrr': 0.10,
    'P@5': 0.10,
    'NDCG@5': 0.10,
    'context_precision': 0.15,
    'context_recall': 0.15,
    'faithfulness': 0.15,
    'answer_relevancy': 0.10,
    'answer_correctness': 0.15,
}

# ─── PESOS DEL RANKING DE GENERACIÓN (A/B/C/D, solo respuestas) ───
# Renormalizado a sumar 1.00 sobre solo las 3 métricas RAGAS de generación,
# manteniendo la proporción relativa del ranking principal:
#   faithfulness: 0.15 → 0.375
#   answer_relevancy: 0.10 → 0.250
#   answer_correctness: 0.15 → 0.375
PESOS_GENERACION = {
    'faithfulness': 0.375,
    'answer_relevancy': 0.250,
    'answer_correctness': 0.375,
}

# Estrategias
TIPOS = ['Directa', 'Caso', 'Compleja', 'Negativa']
CHUNKINGS = ['A_fixed', 'B_markdown', 'C_semantica', 'D_jerarquico']
CHUNKINGS_PRINCIPAL = ['A_fixed', 'B_markdown', 'C_semantica']  # excluye D
RETRIEVALS = ['bm25', 'semantica_generalista', 'semantica_medico',
              'semantica_experto', 'hibrida_generalista', 'hibrida_medico',
              'hibrida_experto']

# ═══════════════════════════════════════════════════════════════════
# CARGA DE DATOS
# ═══════════════════════════════════════════════════════════════════

print("═" * 70)
print(" CONSOLIDACIÓN FINAL DE RESULTADOS")
print("═" * 70)

DIR_GRAFICAS.mkdir(exist_ok=True)


def cargar_json(path):
    if not os.path.exists(path):
        print(f" - [WARN] No encontrado: {path}")
        return []
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


datos_pipeline = cargar_json(ARCHIVO_PIPELINE)
datos_ir = cargar_json(ARCHIVO_IR)
datos_ragas = cargar_json(ARCHIVO_RAGAS)
datos_negativas = cargar_json(ARCHIVO_NEGATIVAS)

print(f" - Pipeline:     {len(datos_pipeline)} ejecuciones")
print(f" - IR:           {len(datos_ir)} ejecuciones")
print(f" - RAGAS:        {len(datos_ragas)} ejecuciones")
print(f" - Negativas:    {len(datos_negativas)} ejecuciones")

df_pipeline = pd.DataFrame(datos_pipeline)
df_ir = pd.DataFrame(datos_ir)
df_ragas = pd.DataFrame(datos_ragas)
df_negativas = pd.DataFrame(datos_negativas) if datos_negativas else pd.DataFrame()


# ═══════════════════════════════════════════════════════════════════
# DIAGNÓSTICO DE CALIDAD DE DATOS RAGAS
# Detecta fallos silenciosos de la API juez (NaN, 0.0 sospechosos)
# ═══════════════════════════════════════════════════════════════════

def diagnosticar_calidad_ragas(df, umbral_alerta_pct=5.0):
    """
    Audita la calidad de los datos RAGAS detectando posibles fallos
    silenciosos del LLM juez (NaN o ceros exactos sospechosos).

    Para D_jerarquico, las métricas de retrieval (context_precision/recall)
    son NaN por diseño y se excluyen del diagnóstico de esas métricas.
    """
    print("\n" + "═" * 70)
    print(" DIAGNÓSTICO DE CALIDAD DE DATOS RAGAS")
    print("═" * 70)

    if df.empty:
        print(" - [WARN] No hay datos RAGAS que diagnosticar.")
        return {}

    metricas_ragas = ['context_precision', 'context_recall',
                      'faithfulness', 'answer_relevancy', 'answer_correctness']
    metricas_existentes = [m for m in metricas_ragas if m in df.columns]

    diagnostico = {}
    hay_alertas = False

    print(f" - Total ejecuciones evaluadas: {len(df)}\n")
    print(f" {'Métrica':<22} {'NaN':>8} {'%NaN':>7} {'==0':>8} {'%==0':>7} {'Válidas':>9} {'Media':>8} {'Mediana':>8}")
    print(" " + "-" * 88)

    for metrica in metricas_existentes:
        # Para D_jerarquico, context_precision y context_recall son NaN por diseño
        if metrica in ('context_precision', 'context_recall'):
            df_metrica = df[df['chunking'] != 'D_jerarquico']
        else:
            df_metrica = df

        n_total = len(df_metrica)
        if n_total == 0:
            continue

        valores = df_metrica[metrica]
        n_nan = valores.isna().sum()
        n_cero = (valores == 0.0).sum()
        n_validas = n_total - n_nan
        media = valores.mean()
        mediana = valores.median()

        pct_nan = (n_nan / n_total) * 100
        pct_cero = (n_cero / n_total) * 100

        diagnostico[metrica] = {
            'n_total': n_total,
            'n_nan': int(n_nan),
            'n_cero_exacto': int(n_cero),
            'n_validas': int(n_validas),
            'pct_nan': round(pct_nan, 2),
            'pct_cero': round(pct_cero, 2),
            'media': round(media, 4) if pd.notna(media) else None,
            'mediana': round(mediana, 4) if pd.notna(mediana) else None,
        }

        marca_nan = ' *' if pct_nan > umbral_alerta_pct else ''
        marca_cero = ' *' if pct_cero > 15.0 else ''
        if marca_nan or marca_cero:
            hay_alertas = True

        print(f" {metrica:<22} {n_nan:>8} {pct_nan:>6.1f}%{marca_nan:<2}"
              f"{n_cero:>7} {pct_cero:>6.1f}%{marca_cero:<2}"
              f"{n_validas:>9} {media:>8.4f} {mediana:>8.4f}")

    print()
    if hay_alertas:
        print(" [!! ALERTA] Algunas métricas superan los umbrales razonables.")
        print("    Las marcadas con * pueden indicar fallos del LLM juez.")
        print(f"    Umbrales: NaN > {umbral_alerta_pct}% | ceros exactos > 15%")
        print()
        print("    ACCIÓN RECOMENDADA antes de usar las métricas finales:")
        print("    1. Revisar manualmente algunas ejecuciones con NaN o 0.0")
        print("    2. Si son fallos reales del juez, reejecutar el script 3")
        print("       borrando solo esas entradas del JSON de checkpoint")
        print("    3. Si son resultados legítimos (sistema RAG malo en esa")
        print("       combinación), continuar normalmente")
    else:
        print(" [OK] Calidad de datos RAGAS dentro de umbrales razonables.")

    if hay_alertas:
        print("\n - Top 5 combinaciones con más NaN o ceros (para investigar):")
        # IMPORTANTE: para D_jerarquico, context_precision y context_recall son
        # NaN por diseño, no por fallo del juez. Hay que contarlos aparte para
        # no saturar el Top 5 con ruido esperado.
        df_aux = df.copy()
        metricas_solo_gen = [m for m in metricas_existentes
                             if m not in ('context_precision', 'context_recall')]

        # Para A/B/C: cuentan TODAS las métricas
        # Para D: cuentan solo las de generación (las otras son NaN por diseño)
        def _contar_problemas(row):
            if row['chunking'] == 'D_jerarquico':
                cols = metricas_solo_gen
            else:
                cols = metricas_existentes
            n_nan = sum(1 for c in cols if pd.isna(row.get(c)))
            n_cero = sum(1 for c in cols if row.get(c) == 0.0)
            return n_nan + n_cero

        df_aux['_problema'] = df_aux.apply(_contar_problemas, axis=1)
        peores = (df_aux.groupby(['chunking', 'retrieval'])['_problema']
                  .sum().sort_values(ascending=False).head(5))
        for (ck, ret), n in peores.items():
            if n > 0:
                print(f"     {ck:>14} + {ret:<22} → {n} valores problemáticos")

    print("═" * 70)
    return diagnostico


diagnostico_ragas = diagnosticar_calidad_ragas(df_ragas)

with open('diagnostico_calidad_ragas.json', 'w', encoding='utf-8') as f:
    json.dump(diagnostico_ragas, f, ensure_ascii=False, indent=2)
print(" - Diagnóstico guardado en: diagnostico_calidad_ragas.json")


# ═══════════════════════════════════════════════════════════════════
# TABLA MAESTRA: TODAS LAS MÉTRICAS POR (chunking × retrieval)
# ═══════════════════════════════════════════════════════════════════

def construir_tabla_maestra():
    """Combina IR + RAGAS + eficiencia + negativas en una sola tabla.

    pandas.groupby().mean() ignora NaN, por lo que las medias salen
    sobre los valores válidos disponibles. D_jerarquico tendrá NaN
    en cols_ir y en context_precision/recall (por diseño).
    """
    cols_ir = ['mrr', 'P@3', 'P@5', 'R@3', 'R@5', 'NDCG@3', 'NDCG@5']
    cols_ragas = ['context_precision', 'context_recall',
                  'faithfulness', 'answer_relevancy', 'answer_correctness']
    cols_eff = ['latencia_retrieval_ms', 'latencia_generacion_ms',
                'latencia_total_ms', 'chars_contexto', 'tokens_contexto_aprox']

    # IR (solo A/B/C: D_jerarquico no se evalúa con métricas IR clásicas)
    ir_agg = (df_ir.groupby(['chunking', 'retrieval'])[cols_ir]
              .mean().round(4).reset_index())

    # RAGAS (todas las combinaciones, con NaN por diseño en D)
    cols_ragas_existentes = [c for c in cols_ragas if c in df_ragas.columns]
    ragas_agg = (df_ragas.groupby(['chunking', 'retrieval'])[cols_ragas_existentes]
                 .mean().round(4).reset_index())

    # Eficiencia (todas)
    eff_agg = (df_pipeline.groupby(['chunking', 'retrieval'])[cols_eff]
               .mean().round(2).reset_index())

    # Negativas
    if not df_negativas.empty:
        neg_agg = (df_negativas.groupby(['chunking', 'retrieval'])
                   .agg(declination_rate=('declino_correctamente', 'mean'),
                        hallucination_rate=('alucino', 'mean'))
                   .round(4).reset_index())
    else:
        neg_agg = pd.DataFrame(columns=['chunking', 'retrieval'])

    # Merge OUTER para no perder D (que falta en ir_agg)
    tabla = ir_agg.merge(ragas_agg, on=['chunking', 'retrieval'], how='outer')
    tabla = tabla.merge(eff_agg, on=['chunking', 'retrieval'], how='outer')
    if not neg_agg.empty:
        tabla = tabla.merge(neg_agg, on=['chunking', 'retrieval'], how='outer')

    return tabla


tabla_maestra = construir_tabla_maestra()
print(f"\n - Tabla maestra construida: {len(tabla_maestra)} combinaciones")


# ═══════════════════════════════════════════════════════════════════
# RANKINGS — DOS RANKINGS SEPARADOS POR DISEÑO METODOLÓGICO
# ═══════════════════════════════════════════════════════════════════

def calcular_score_compuesto(df, pesos):
    """Score ponderado normalizado.

    fillna(0) penaliza las celdas agregadas que sean NaN. Los pesos
    deben sumar 1.0 sobre el conjunto de métricas que aplican a la
    combinación evaluada.
    """
    df = df.copy()
    componentes = []
    pesos_aplicados = []
    for metrica, peso in pesos.items():
        if metrica in df.columns:
            valores = df[metrica].fillna(0)
            componentes.append(valores * peso)
            pesos_aplicados.append((metrica, peso))
    df['score_compuesto'] = sum(componentes).round(4) if componentes else 0
    return df.sort_values('score_compuesto', ascending=False).reset_index(drop=True)


# ─── Ranking PRINCIPAL: A/B/C × 7, todas las métricas ───
print("\n" + "═" * 70)
print(" RANKING PRINCIPAL (A/B/C end-to-end)")
print("═" * 70)
print(f" - Métricas usadas: {list(PESOS_PRINCIPAL.keys())}")
print(f" - Suma de pesos:   {sum(PESOS_PRINCIPAL.values()):.4f}")

tabla_principal = tabla_maestra[tabla_maestra['chunking'].isin(CHUNKINGS_PRINCIPAL)].copy()
ranking_principal = calcular_score_compuesto(tabla_principal, PESOS_PRINCIPAL)

print(f"\n - Top 5 combinaciones (ranking principal):")
for i, row in ranking_principal.head(5).iterrows():
    print(f"     {i+1}. {row['chunking']:>12} + {row['retrieval']:<22} → {row['score_compuesto']:.4f}")


# ─── Ranking GENERACIÓN: A/B/C/D × 7, solo métricas RAGAS de generación ───
print("\n" + "═" * 70)
print(" RANKING DE GENERACIÓN (A/B/C/D, solo métricas RAGAS de respuesta)")
print("═" * 70)
print(f" - Métricas usadas: {list(PESOS_GENERACION.keys())}")
print(f" - Suma de pesos:   {sum(PESOS_GENERACION.values()):.4f}")
print(" - Permite incluir D_jerarquico para responder:")
print("   ¿qué chunking produce las mejores RESPUESTAS finales,")
print("    independientemente de cómo se recuperaron los chunks?")

ranking_generacion = calcular_score_compuesto(tabla_maestra.copy(), PESOS_GENERACION)
ranking_generacion = ranking_generacion.rename(columns={'score_compuesto': 'score_generacion'})

print(f"\n - Top 5 combinaciones (ranking de generación):")
for i, row in ranking_generacion.head(5).iterrows():
    print(f"     {i+1}. {row['chunking']:>12} + {row['retrieval']:<22} → {row['score_generacion']:.4f}")

# Posición de D_jerarquico en el ranking de generación
d_rows = ranking_generacion[ranking_generacion['chunking'] == 'D_jerarquico']
if not d_rows.empty:
    mejor_d_idx = d_rows.index.min()
    mejor_d = d_rows.loc[mejor_d_idx]
    print(f"\n - Mejor combinación con D_jerarquico:")
    print(f"     Posición #{mejor_d_idx + 1}: {mejor_d['chunking']} + {mejor_d['retrieval']}")
    print(f"     score_generacion = {mejor_d['score_generacion']:.4f}")


# ═══════════════════════════════════════════════════════════════════
# AGREGACIONES POR DIMENSIÓN
# ═══════════════════════════════════════════════════════════════════

def agregar_por(df, agrupar_por, columnas):
    cols_existentes = [c for c in columnas if c in df.columns]
    return (df.groupby(agrupar_por)[cols_existentes]
            .mean().round(4).reset_index())


# Agregaciones IR (solo A/B/C disponibles)
cols_ir = ['mrr', 'P@3', 'P@5', 'R@3', 'R@5', 'NDCG@3', 'NDCG@5']
ir_por_chunking = agregar_por(df_ir, 'chunking', cols_ir)
ir_por_retrieval = agregar_por(df_ir, 'retrieval', cols_ir)
ir_por_tipo = agregar_por(df_ir, ['chunking', 'retrieval', 'tipo'], cols_ir)

# Agregaciones RAGAS (todas las combinaciones; D tendrá NaN en cp/cr)
cols_ragas = ['context_precision', 'context_recall',
              'faithfulness', 'answer_relevancy', 'answer_correctness']
ragas_por_chunking = agregar_por(df_ragas, 'chunking', cols_ragas)
ragas_por_retrieval = agregar_por(df_ragas, 'retrieval', cols_ragas)
ragas_por_tipo = agregar_por(df_ragas, ['chunking', 'retrieval', 'tipo'], cols_ragas)

# Agregaciones eficiencia
cols_eff = ['latencia_retrieval_ms', 'latencia_generacion_ms',
            'latencia_total_ms', 'chars_contexto', 'tokens_contexto_aprox']
eff_por_chunking = (df_pipeline.groupby('chunking')[cols_eff].mean().round(2).reset_index())
eff_por_retrieval = (df_pipeline.groupby('retrieval')[cols_eff].mean().round(2).reset_index())


# ═══════════════════════════════════════════════════════════════════
# GRÁFICOS
# ═══════════════════════════════════════════════════════════════════

def heatmap_metrica(df, metrica, titulo, archivo, cmap='RdYlGn'):
    """Heatmap de una métrica con chunking × retrieval."""
    if metrica not in df.columns:
        return
    pivot = df.pivot_table(index='chunking', columns='retrieval',
                            values=metrica, aggfunc='mean')
    pivot = pivot.reindex(index=[c for c in CHUNKINGS if c in pivot.index],
                          columns=[r for r in RETRIEVALS if r in pivot.columns])

    fig, ax = plt.subplots(figsize=(10, 4.5))
    sns.heatmap(pivot, annot=True, fmt='.3f', cmap=cmap, ax=ax,
                cbar_kws={'label': metrica}, linewidths=0.5,
                vmin=0, vmax=1 if pivot.max().max() <= 1 else None)
    ax.set_title(titulo, fontweight='bold')
    ax.set_xlabel('Estrategia de retrieval')
    ax.set_ylabel('Estrategia de chunking')
    plt.xticks(rotation=30, ha='right')
    plt.tight_layout()
    plt.savefig(DIR_GRAFICAS / archivo)
    plt.close()


def barras_por_dimension(df, columna_x, metricas, titulo, archivo, paleta):
    """Barras agrupadas mostrando varias métricas para cada estrategia."""
    metricas_existentes = [m for m in metricas if m in df.columns]
    if not metricas_existentes:
        return

    df_plot = df[[columna_x] + metricas_existentes].set_index(columna_x)
    fig, ax = plt.subplots(figsize=(11, 5.5))
    df_plot.plot.bar(ax=ax, edgecolor='white', linewidth=0.5)
    ax.set_title(titulo, fontweight='bold')
    ax.set_ylabel('Valor de métrica')
    ax.set_xlabel('')
    ax.legend(loc='upper left', bbox_to_anchor=(1, 1), title='Métrica')
    plt.xticks(rotation=20, ha='right')
    ax.set_ylim(0, max(1.0, df_plot.max().max() * 1.1))
    plt.tight_layout()
    plt.savefig(DIR_GRAFICAS / archivo)
    plt.close()


def scatter_calidad_vs_latencia(tabla, score_col, archivo, titulo_extra=''):
    """Pareto: calidad (score) vs latencia total."""
    if score_col not in tabla.columns or 'latencia_total_ms' not in tabla.columns:
        return

    fig, ax = plt.subplots(figsize=(10, 6))
    for chunking in tabla['chunking'].unique():
        subset = tabla[tabla['chunking'] == chunking]
        ax.scatter(subset['latencia_total_ms'], subset[score_col],
                   s=120, alpha=0.75, label=chunking,
                   color=PALETA_CHUNKING.get(chunking, 'gray'),
                   edgecolors='black', linewidth=0.5)
        for _, row in subset.iterrows():
            ax.annotate(row['retrieval'].replace('semantica_', 's_').replace('hibrida_', 'h_'),
                        (row['latencia_total_ms'], row[score_col]),
                        fontsize=7, alpha=0.7,
                        xytext=(5, 5), textcoords='offset points')

    ax.set_xlabel('Latencia total (ms) — log scale')
    ax.set_ylabel(f'Score ({score_col})')
    titulo_base = 'Trade-off Calidad vs Eficiencia Computacional'
    ax.set_title(f'{titulo_base}{titulo_extra}', fontweight='bold')
    ax.set_xscale('log')
    ax.legend(title='Chunking', loc='best')
    plt.tight_layout()
    plt.savefig(DIR_GRAFICAS / archivo)
    plt.close()


def radar_top_n(tabla, metricas_radar, score_col, n=3, archivo='radar_top_n.png'):
    """Radar chart comparando las top-N combinaciones."""
    metricas_existentes = [m for m in metricas_radar if m in tabla.columns]
    if len(metricas_existentes) < 3:
        return

    top = tabla.head(n)
    angulos = np.linspace(0, 2 * np.pi, len(metricas_existentes), endpoint=False).tolist()
    angulos += angulos[:1]

    fig, ax = plt.subplots(figsize=(8, 8), subplot_kw=dict(polar=True))
    cmap = plt.cm.tab10

    for i, (_, row) in enumerate(top.iterrows()):
        valores = [row[m] if pd.notna(row[m]) else 0 for m in metricas_existentes]
        valores += valores[:1]
        etiqueta = f"{row['chunking']} + {row['retrieval'].replace('semantica_', 's_').replace('hibrida_', 'h_')}"
        ax.plot(angulos, valores, linewidth=2, label=etiqueta, color=cmap(i))
        ax.fill(angulos, valores, alpha=0.15, color=cmap(i))

    ax.set_xticks(angulos[:-1])
    ax.set_xticklabels(metricas_existentes, size=9)
    ax.set_ylim(0, 1)
    ax.set_yticks([0.2, 0.4, 0.6, 0.8, 1.0])
    ax.set_yticklabels(['0.2', '0.4', '0.6', '0.8', '1.0'], size=8)
    ax.set_title(f'Comparativa Top-{n} combinaciones', fontweight='bold', pad=20)
    ax.legend(loc='upper right', bbox_to_anchor=(1.3, 1.1), fontsize=9)
    ax.grid(True)
    plt.tight_layout()
    plt.savefig(DIR_GRAFICAS / archivo)
    plt.close()


def barras_por_tipo_pregunta(df, metricas, titulo, archivo):
    """Compara métricas desglosadas por tipo de pregunta."""
    metricas_existentes = [m for m in metricas if m in df.columns]
    if not metricas_existentes or 'tipo' not in df.columns:
        return

    agg = df.groupby('tipo')[metricas_existentes].mean().round(4)
    if agg.empty:
        return
    agg = agg.reindex([t for t in TIPOS if t in agg.index])

    fig, ax = plt.subplots(figsize=(10, 5))
    agg.plot.bar(ax=ax, edgecolor='white', linewidth=0.5)
    ax.set_title(titulo, fontweight='bold')
    ax.set_ylabel('Valor de métrica')
    ax.set_xlabel('Tipo de pregunta')
    ax.legend(loc='upper left', bbox_to_anchor=(1, 1), title='Métrica')
    plt.xticks(rotation=0)
    ax.set_ylim(0, max(1.0, agg.max().max() * 1.1))
    plt.tight_layout()
    plt.savefig(DIR_GRAFICAS / archivo)
    plt.close()


def comparativa_d_vs_resto(archivo='comparativa_d_vs_resto.png'):
    """D vs A/B/C en métricas de generación + eficiencia."""
    if df_ragas.empty:
        return

    cols_gen = ['faithfulness', 'answer_relevancy', 'answer_correctness']
    cols_existentes = [c for c in cols_gen if c in df_ragas.columns]
    if not cols_existentes:
        return

    df_plot = df_ragas.groupby('chunking')[cols_existentes].mean().round(4)
    df_plot = df_plot.reindex([c for c in CHUNKINGS if c in df_plot.index])

    fig, ax = plt.subplots(figsize=(10, 5.5))
    df_plot.plot.bar(ax=ax, edgecolor='white', linewidth=0.5,
                     color=['#2E86C1', '#28B463', '#E67E22'])
    ax.set_title('Métricas de Generación: estrategia D (jerárquica) vs A/B/C',
                 fontweight='bold')
    ax.set_ylabel('Valor de métrica')
    ax.set_xlabel('')
    ax.legend(loc='lower right', title='Métrica')
    plt.xticks(rotation=0)
    ax.set_ylim(0, 1.05)
    plt.tight_layout()
    plt.savefig(DIR_GRAFICAS / archivo)
    plt.close()


def trade_off_d_vs_resto(archivo='tradeoff_d_calidad_vs_tokens.png'):
    """Trade-off específico de D: ganancia en calidad vs coste en tokens."""
    if df_ragas.empty or df_pipeline.empty:
        return

    cols_gen = ['faithfulness', 'answer_relevancy', 'answer_correctness']
    cols_existentes = [c for c in cols_gen if c in df_ragas.columns]
    if not cols_existentes:
        return

    # Score de generación medio por chunking
    score_gen_chunking = df_ragas.groupby('chunking')[cols_existentes].mean().mean(axis=1)
    tokens_chunking = df_pipeline.groupby('chunking')['tokens_contexto_aprox'].mean()

    df_plot = pd.DataFrame({
        'score_generacion_medio': score_gen_chunking,
        'tokens_contexto_medio': tokens_chunking,
    }).dropna()

    fig, ax = plt.subplots(figsize=(9, 6))
    for chunking in df_plot.index:
        ax.scatter(df_plot.loc[chunking, 'tokens_contexto_medio'],
                   df_plot.loc[chunking, 'score_generacion_medio'],
                   s=350, alpha=0.85, label=chunking,
                   color=PALETA_CHUNKING.get(chunking, 'gray'),
                   edgecolors='black', linewidth=1.0)
        ax.annotate(chunking,
                    (df_plot.loc[chunking, 'tokens_contexto_medio'],
                     df_plot.loc[chunking, 'score_generacion_medio']),
                    fontsize=10, fontweight='bold',
                    xytext=(10, 10), textcoords='offset points')

    ax.set_xlabel('Tokens de contexto medio (proxy de coste)')
    ax.set_ylabel('Score de generación medio (Faith + AR + AC)/3')
    ax.set_title('Trade-off por chunking: calidad de respuesta vs coste de contexto',
                 fontweight='bold')
    plt.tight_layout()
    plt.savefig(DIR_GRAFICAS / archivo)
    plt.close()


def negativas_chart(archivo='negativas_declination_rate.png'):
    """Tasa de declinación correcta para Negativas."""
    if df_negativas.empty:
        return

    agg = (df_negativas.groupby(['chunking', 'retrieval'])['declino_correctamente']
           .mean().round(4).reset_index())
    pivot = agg.pivot_table(index='chunking', columns='retrieval',
                             values='declino_correctamente')
    pivot = pivot.reindex(index=[c for c in CHUNKINGS if c in pivot.index],
                          columns=[r for r in RETRIEVALS if r in pivot.columns])

    fig, ax = plt.subplots(figsize=(10, 4.5))
    sns.heatmap(pivot, annot=True, fmt='.3f', cmap='RdYlGn', ax=ax,
                cbar_kws={'label': 'Tasa de declinación'},
                linewidths=0.5, vmin=0, vmax=1)
    ax.set_title('Robustez ante preguntas Negativas: tasa de declinación correcta',
                 fontweight='bold')
    ax.set_xlabel('Estrategia de retrieval')
    ax.set_ylabel('Estrategia de chunking')
    plt.xticks(rotation=30, ha='right')
    plt.tight_layout()
    plt.savefig(DIR_GRAFICAS / archivo)
    plt.close()


# ═══════════════════════════════════════════════════════════════════
# GENERAR TODAS LAS GRÁFICAS
# ═══════════════════════════════════════════════════════════════════

print("\n - Generando gráficas...")

# Heatmaps de métricas clave (incluyen D_jerarquico cuando aplica)
heatmap_metrica(df_ir, 'mrr', 'MRR por combinación (IR clásico)',
                'heatmap_mrr.png')
heatmap_metrica(df_ir, 'NDCG@5', 'NDCG@5 por combinación (IR clásico)',
                'heatmap_ndcg5.png')
heatmap_metrica(df_ragas, 'context_precision',
                'Context Precision por combinación (RAGAS)',
                'heatmap_context_precision.png')
heatmap_metrica(df_ragas, 'context_recall',
                'Context Recall por combinación (RAGAS)',
                'heatmap_context_recall.png')
heatmap_metrica(df_ragas, 'faithfulness',
                'Faithfulness por combinación (RAGAS)',
                'heatmap_faithfulness.png')
heatmap_metrica(df_ragas, 'answer_correctness',
                'Answer Correctness por combinación (RAGAS)',
                'heatmap_answer_correctness.png')

# Comparativas agregadas
barras_por_dimension(ir_por_chunking, 'chunking',
                     ['mrr', 'P@5', 'R@5', 'NDCG@5'],
                     'Métricas IR agregadas por estrategia de chunking',
                     'barras_ir_por_chunking.png', PALETA_CHUNKING)

barras_por_dimension(ir_por_retrieval, 'retrieval',
                     ['mrr', 'P@5', 'R@5', 'NDCG@5'],
                     'Métricas IR agregadas por estrategia de retrieval',
                     'barras_ir_por_retrieval.png', PALETA_RETRIEVAL)

barras_por_dimension(ragas_por_chunking, 'chunking',
                     cols_ragas,
                     'Métricas RAGAS agregadas por estrategia de chunking',
                     'barras_ragas_por_chunking.png', PALETA_CHUNKING)

barras_por_dimension(ragas_por_retrieval, 'retrieval',
                     cols_ragas,
                     'Métricas RAGAS agregadas por estrategia de retrieval',
                     'barras_ragas_por_retrieval.png', PALETA_RETRIEVAL)

# Trade-offs (uno por cada ranking)
scatter_calidad_vs_latencia(ranking_principal, 'score_compuesto',
                            'tradeoff_principal.png',
                            ' (ranking principal A/B/C)')
scatter_calidad_vs_latencia(ranking_generacion, 'score_generacion',
                            'tradeoff_generacion.png',
                            ' (ranking generación A/B/C/D)')

# Trade-off específico de D
trade_off_d_vs_resto()

# Radar top-3 (uno por cada ranking)
radar_top_n(ranking_principal,
            ['mrr', 'NDCG@5', 'context_precision', 'context_recall',
             'faithfulness', 'answer_relevancy', 'answer_correctness'],
            'score_compuesto', n=3, archivo='radar_top3_principal.png')
radar_top_n(ranking_generacion,
            ['faithfulness', 'answer_relevancy', 'answer_correctness'],
            'score_generacion', n=3, archivo='radar_top3_generacion.png')

# Por tipo de pregunta
barras_por_tipo_pregunta(df_ir, ['mrr', 'P@5', 'NDCG@5'],
                          'Métricas IR por tipo de pregunta',
                          'tipo_pregunta_ir.png')
barras_por_tipo_pregunta(df_ragas, cols_ragas,
                          'Métricas RAGAS por tipo de pregunta',
                          'tipo_pregunta_ragas.png')

# D vs A/B/C
comparativa_d_vs_resto()

# Negativas
negativas_chart()

print(f"   Gráficas guardadas en: {DIR_GRAFICAS}/")


# ═══════════════════════════════════════════════════════════════════
# EXCEL FINAL CON FORMATO
# ═══════════════════════════════════════════════════════════════════

print("\n - Generando Excel consolidado...")

wb = openpyxl.Workbook()
wb.remove(wb.active)

thin = Side(style='thin', color='BBBBBB')
borde = Border(left=thin, right=thin, top=thin, bottom=thin)
COLOR_TITULO = '1F3864'
COLOR_HEADER = 'CCCCCC'
COLOR_TOTAL = 'EEEEEE'


def estilo(cell, bold=False, bg=None, color='000000', size=10):
    cell.font = Font(name='Arial', bold=bold, color=color, size=size)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = borde
    if bg:
        cell.fill = PatternFill('solid', start_color=bg)


def añadir_hoja(nombre, df_h, titulo, anchos=None, formato_color=True):
    ws = wb.create_sheet(nombre)
    n_cols = len(df_h.columns)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    estilo(ws.cell(row=1, column=1, value=titulo), bold=True,
           bg=COLOR_TITULO, color='FFFFFF', size=11)
    ws.row_dimensions[1].height = 26

    for col, cab in enumerate(df_h.columns, 1):
        estilo(ws.cell(row=2, column=col, value=str(cab)), bold=True,
               bg=COLOR_HEADER)
        ws.column_dimensions[get_column_letter(col)].width = (
            anchos[col - 1] if anchos and col - 1 < len(anchos) else 14)

    for fila, (_, row) in enumerate(df_h.iterrows(), 3):
        for col, val in enumerate(row, 1):
            if isinstance(val, float):
                val = round(val, 4) if pd.notna(val) else '-'
            estilo(ws.cell(row=fila, column=col, value=val))

    if formato_color and len(df_h) > 1:
        for col_idx, col_nombre in enumerate(df_h.columns, 1):
            valores = df_h[col_nombre]
            if pd.api.types.is_numeric_dtype(valores) and valores.notna().any():
                col_letter = get_column_letter(col_idx)
                rango = f"{col_letter}3:{col_letter}{len(df_h) + 2}"
                if 'latencia' in col_nombre or 'chars' in col_nombre or 'tokens' in col_nombre:
                    rule = ColorScaleRule(
                        start_type='min', start_color='63BE7B',
                        mid_type='percentile', mid_value=50, mid_color='FFEB84',
                        end_type='max', end_color='F8696B'
                    )
                else:
                    rule = ColorScaleRule(
                        start_type='min', start_color='F8696B',
                        mid_type='percentile', mid_value=50, mid_color='FFEB84',
                        end_type='max', end_color='63BE7B'
                    )
                ws.conditional_formatting.add(rango, rule)


# ─── Hojas: tabla maestra y dos rankings ───
añadir_hoja('Tabla Maestra', tabla_maestra,
            'TABLA MAESTRA — TODAS LAS MÉTRICAS POR COMBINACIÓN')

añadir_hoja('Ranking Principal', ranking_principal,
            'RANKING PRINCIPAL — A/B/C END-TO-END (todas las métricas)')

añadir_hoja('Ranking Generación', ranking_generacion,
            'RANKING DE GENERACIÓN — A/B/C/D (solo Faithfulness + AR + AC)')

# IR
añadir_hoja('IR Global', agregar_por(df_ir, ['chunking', 'retrieval'], cols_ir),
            'MÉTRICAS IR — VISTA GLOBAL (excluye D_jerarquico)')
añadir_hoja('IR Por Chunking', ir_por_chunking,
            'MÉTRICAS IR — AGREGADAS POR CHUNKING')
añadir_hoja('IR Por Retrieval', ir_por_retrieval,
            'MÉTRICAS IR — AGREGADAS POR RETRIEVAL')
añadir_hoja('IR Por Tipo', ir_por_tipo,
            'MÉTRICAS IR — DESGLOSE POR TIPO DE PREGUNTA')

# RAGAS
añadir_hoja('RAGAS Global',
            agregar_por(df_ragas, ['chunking', 'retrieval'], cols_ragas),
            'MÉTRICAS RAGAS — VISTA GLOBAL')
añadir_hoja('RAGAS Por Chunking', ragas_por_chunking,
            'MÉTRICAS RAGAS — AGREGADAS POR CHUNKING')
añadir_hoja('RAGAS Por Retrieval', ragas_por_retrieval,
            'MÉTRICAS RAGAS — AGREGADAS POR RETRIEVAL')
añadir_hoja('RAGAS Por Tipo', ragas_por_tipo,
            'MÉTRICAS RAGAS — DESGLOSE POR TIPO DE PREGUNTA')

# Eficiencia
añadir_hoja('Eficiencia Por Chunking', eff_por_chunking,
            'EFICIENCIA COMPUTACIONAL — POR CHUNKING')
añadir_hoja('Eficiencia Por Retrieval', eff_por_retrieval,
            'EFICIENCIA COMPUTACIONAL — POR RETRIEVAL')

# Negativas
if not df_negativas.empty:
    neg_global = (df_negativas.groupby(['chunking', 'retrieval'])
                  .agg(declination_rate=('declino_correctamente', 'mean'),
                       hallucination_rate=('alucino', 'mean'),
                       n_negativas=('declino_correctamente', 'count'))
                  .round(4).reset_index())
    añadir_hoja('Negativas Global', neg_global,
                'NEGATIVAS — VISTA GLOBAL')

# Diagnóstico de calidad
if diagnostico_ragas:
    diag_df = pd.DataFrame(diagnostico_ragas).T.reset_index()
    diag_df = diag_df.rename(columns={'index': 'metrica'})
    añadir_hoja('Diagnóstico Calidad', diag_df,
                'DIAGNÓSTICO DE CALIDAD DE DATOS RAGAS (NaN, ceros)',
                formato_color=False)

wb.save(ARCHIVO_EXCEL)
print(f"   Excel guardado: {ARCHIVO_EXCEL}")


# ═══════════════════════════════════════════════════════════════════
# RESUMEN EJECUTIVO EN MARKDOWN
# ═══════════════════════════════════════════════════════════════════

def generar_resumen():
    lineas = []
    lineas.append('# Resumen Ejecutivo: Evaluación del Sistema RAG\n')

    # ─── RANKING PRINCIPAL ───
    lineas.append('## 1. Ranking principal (end-to-end)\n')
    lineas.append('Compara las **21 combinaciones A/B/C × 7 retrievals** con todas '
                  'las métricas (IR clásicas + RAGAS retrieval + RAGAS generación).\n')
    lineas.append('La estrategia D_jerarquica se excluye de este ranking porque sus '
                  'métricas de retrieval son `null` por diseño (los chunks padre '
                  'rompen la direccionalidad de la coincidencia fuzzy contra la '
                  'macro-sección, inflando artificialmente Context Precision/Recall).\n')

    if not ranking_principal.empty:
        top1 = ranking_principal.iloc[0]
        lineas.append(f"### Combinación ganadora\n")
        lineas.append(f"**`{top1['chunking']}` + `{top1['retrieval']}`** "
                      f"con score compuesto = **{top1['score_compuesto']:.4f}**\n")
        lineas.append('\n**Desglose de métricas:**\n')
        metricas_top = [
            ('MRR', 'mrr'),
            ('Precision@5', 'P@5'),
            ('Recall@5', 'R@5'),
            ('NDCG@5', 'NDCG@5'),
            ('Context Precision', 'context_precision'),
            ('Context Recall', 'context_recall'),
            ('Faithfulness', 'faithfulness'),
            ('Answer Relevancy', 'answer_relevancy'),
            ('Answer Correctness', 'answer_correctness'),
        ]
        for nombre, col in metricas_top:
            if col in top1.index and pd.notna(top1[col]):
                lineas.append(f"- **{nombre}**: {top1[col]:.4f}")
        lineas.append('')

        if pd.notna(top1.get('latencia_total_ms')):
            lineas.append(f"**Latencia total media**: {top1['latencia_total_ms']:.0f} ms\n")

        lineas.append('### Top 5 (ranking principal)\n')
        lineas.append('| # | Chunking | Retrieval | Score |')
        lineas.append('|---|----------|-----------|-------|')
        for i, row in ranking_principal.head(5).iterrows():
            lineas.append(f"| {i+1} | `{row['chunking']}` | `{row['retrieval']}` | "
                          f"{row['score_compuesto']:.4f} |")
        lineas.append('')

    # ─── RANKING GENERACIÓN ───
    lineas.append('## 2. Ranking de generación (incluye D_jerárquico)\n')
    lineas.append('Compara las **28 combinaciones A/B/C/D × 7 retrievals** únicamente '
                  'con métricas RAGAS de generación (Faithfulness, Answer Relevancy, '
                  'Answer Correctness). Pesos renormalizados a sumar 1:\n')
    lineas.append('```')
    for k, v in PESOS_GENERACION.items():
        lineas.append(f"  {k:<22} → peso {v:.3f}")
    lineas.append('```\n')
    lineas.append('Este ranking responde a la pregunta secundaria: **¿qué chunking '
                  'produce las mejores respuestas, independientemente de cómo se '
                  'recuperaron los chunks?** Permite valorar si el sobrecoste de '
                  'contexto de D_jerárquico se compensa con mejor calidad de respuesta.\n')

    if not ranking_generacion.empty:
        top1_g = ranking_generacion.iloc[0]
        lineas.append(f"### Combinación ganadora (generación)\n")
        lineas.append(f"**`{top1_g['chunking']}` + `{top1_g['retrieval']}`** "
                      f"con score_generacion = **{top1_g['score_generacion']:.4f}**\n")

        lineas.append('### Top 5 (ranking generación)\n')
        lineas.append('| # | Chunking | Retrieval | Score Gen | Faith | AR | AC |')
        lineas.append('|---|----------|-----------|-----------|-------|-----|-----|')
        for i, row in ranking_generacion.head(5).iterrows():
            f_val = f"{row['faithfulness']:.4f}" if pd.notna(row.get('faithfulness')) else '-'
            ar_val = f"{row['answer_relevancy']:.4f}" if pd.notna(row.get('answer_relevancy')) else '-'
            ac_val = f"{row['answer_correctness']:.4f}" if pd.notna(row.get('answer_correctness')) else '-'
            lineas.append(f"| {i+1} | `{row['chunking']}` | `{row['retrieval']}` | "
                          f"{row['score_generacion']:.4f} | {f_val} | {ar_val} | {ac_val} |")
        lineas.append('')

        # Posición de D
        d_rows = ranking_generacion[ranking_generacion['chunking'] == 'D_jerarquico']
        if not d_rows.empty:
            mejor_d_idx = d_rows.index.min()
            mejor_d = d_rows.loc[mejor_d_idx]
            lineas.append(f"### Mejor combinación con D_jerárquico\n")
            lineas.append(f"Posición **#{mejor_d_idx + 1}** del ranking de generación: "
                          f"`{mejor_d['chunking']}` + `{mejor_d['retrieval']}` "
                          f"(score_generacion = {mejor_d['score_generacion']:.4f})\n")
            if pd.notna(mejor_d.get('tokens_contexto_aprox')):
                # Comparar con el mejor de A/B/C en tokens
                no_d = ranking_generacion[ranking_generacion['chunking'] != 'D_jerarquico']
                if not no_d.empty and 'tokens_contexto_aprox' in no_d.columns:
                    tokens_mejor_no_d = no_d.iloc[0].get('tokens_contexto_aprox')
                    if pd.notna(tokens_mejor_no_d) and tokens_mejor_no_d > 0:
                        ratio = mejor_d['tokens_contexto_aprox'] / tokens_mejor_no_d
                        lineas.append(f"- Tokens de contexto medio: **{mejor_d['tokens_contexto_aprox']:.0f}** "
                                      f"(≈ {ratio:.1f}× más que el ganador del ranking de generación sin D)\n")

    # ─── HALLAZGOS POR DIMENSIÓN ───
    lineas.append('## 3. Hallazgos principales\n')

    if not ir_por_chunking.empty and 'mrr' in ir_por_chunking.columns:
        mejor_ck_ir = ir_por_chunking.loc[ir_por_chunking['mrr'].idxmax()]
        lineas.append(f"### Mejor estrategia de chunking (IR clásico)\n")
        lineas.append(f"`{mejor_ck_ir['chunking']}` con MRR = {mejor_ck_ir['mrr']:.4f} "
                      f"y NDCG@5 = {mejor_ck_ir['NDCG@5']:.4f}\n")

    if not ir_por_retrieval.empty and 'mrr' in ir_por_retrieval.columns:
        mejor_ret_ir = ir_por_retrieval.loc[ir_por_retrieval['mrr'].idxmax()]
        lineas.append(f"### Mejor estrategia de retrieval (IR clásico)\n")
        lineas.append(f"`{mejor_ret_ir['retrieval']}` con MRR = {mejor_ret_ir['mrr']:.4f} "
                      f"y NDCG@5 = {mejor_ret_ir['NDCG@5']:.4f}\n")

    if not ragas_por_chunking.empty and 'answer_correctness' in ragas_por_chunking.columns:
        mejor_ck_gen = ragas_por_chunking.loc[ragas_por_chunking['answer_correctness'].idxmax()]
        lineas.append(f"### Mejor estrategia de chunking (calidad de generación)\n")
        lineas.append(f"`{mejor_ck_gen['chunking']}` con Answer Correctness = "
                      f"{mejor_ck_gen['answer_correctness']:.4f}\n")

    if not eff_por_chunking.empty:
        mas_rapido = eff_por_chunking.loc[eff_por_chunking['latencia_total_ms'].idxmin()]
        lineas.append(f"### Estrategia de chunking más rápida\n")
        lineas.append(f"`{mas_rapido['chunking']}` con latencia total media de "
                      f"{mas_rapido['latencia_total_ms']:.0f} ms\n")

    if not df_negativas.empty:
        nar_por_chunking = (df_negativas.groupby('chunking')['declino_correctamente']
                            .mean().round(4))
        mejor_robustez = nar_por_chunking.idxmax()
        lineas.append(f"### Mejor robustez ante preguntas Negativas\n")
        lineas.append(f"`{mejor_robustez}` con tasa de declinación = "
                      f"{nar_por_chunking.max():.4f} ({nar_por_chunking.max()*100:.1f}%)\n")

    if not ragas_por_retrieval.empty and 'answer_correctness' in ragas_por_retrieval.columns:
        lineas.append('### Comparativa de modelos de embeddings (Answer Correctness)\n')
        for tipo_modelo in ['generalista', 'experto', 'medico']:
            row_sem = ragas_por_retrieval[
                ragas_por_retrieval['retrieval'] == f'semantica_{tipo_modelo}']
            if not row_sem.empty:
                lineas.append(f"- Retrieval semántico **{tipo_modelo}**: "
                              f"Answer Correctness = "
                              f"{row_sem.iloc[0]['answer_correctness']:.4f}")
        lineas.append('')

    # ─── DIAGNÓSTICO DE CALIDAD ───
    if diagnostico_ragas:
        lineas.append('## 4. Diagnóstico de calidad de los datos RAGAS\n')
        lineas.append('Auditoría de fallos del LLM juez (NaN o ceros sospechosos):\n')
        lineas.append('| Métrica | NaN | %NaN | Ceros | %Ceros | Válidas |')
        lineas.append('|---------|-----|------|-------|--------|---------|')
        for met, stats in diagnostico_ragas.items():
            lineas.append(f"| `{met}` | {stats['n_nan']} | {stats['pct_nan']}% | "
                          f"{stats['n_cero_exacto']} | {stats['pct_cero']}% | "
                          f"{stats['n_validas']} |")
        lineas.append('')

    # ─── LIMITACIONES ───
    lineas.append('## 5. Limitaciones metodológicas reconocidas\n')
    lineas.append('1. **Estrategia D (jerárquica)** se excluye del ranking principal '
                  'y de las métricas de retrieval (IR clásicas y Context Precision/Recall '
                  'de RAGAS). Sus chunks padre de ~8k caracteres rompen la direccionalidad '
                  'de la coincidencia fuzzy contra la macro-sección, inflando '
                  'artificialmente las métricas de recuperación. Se evalúa por separado '
                  'en el ranking de generación.\n')
    lineas.append('2. **Criterio de relevancia IR clásico** asume que la macro-sección '
                  'de origen define el universo de chunks relevantes para una pregunta. '
                  'Esto es una aproximación: chunks dentro pueden ser tangenciales y '
                  'chunks fuera pueden aportar información útil. El sesgo es sistemático '
                  'y afecta por igual a todas las estrategias comparadas.\n')
    lineas.append('3. **Preguntas Negativas** se evalúan con métricas binarias '
                  'específicas (declination rate vs hallucination rate), no con RAGAS '
                  'estándar, por incompatibilidad conceptual del framework con preguntas '
                  'fuera de alcance.\n')
    lineas.append('4. **Manejo de nulos en pandas**: cuando una métrica concreta de una '
                  'ejecución es NaN (por fallo del LLM juez), pandas la ignora en el '
                  'cálculo de medias por agrupación. El registro completo NO se descarta: '
                  'sus otras métricas válidas siguen contando. Sin embargo, el score '
                  'compuesto aplica `fillna(0)` sobre la tabla agregada como política '
                  'defensiva: si una combinación tiene muchos NaN, su score baja.\n')

    # ─── DEFINICIÓN DE PESOS ───
    lineas.append('## 6. Definición de los scores compuestos\n')

    lineas.append('### Ranking principal (A/B/C, suma = 1.00)\n')
    lineas.append('```')
    for k, v in PESOS_PRINCIPAL.items():
        lineas.append(f"  {k:<22} → peso {v:.2f}")
    lineas.append('```\n')

    lineas.append('### Ranking de generación (A/B/C/D, suma = 1.00)\n')
    lineas.append('```')
    for k, v in PESOS_GENERACION.items():
        lineas.append(f"  {k:<22} → peso {v:.3f}")
    lineas.append('```\n')

    return '\n'.join(lineas)


print("\n - Generando resumen ejecutivo...")
resumen_md = generar_resumen()
with open(ARCHIVO_RESUMEN, 'w', encoding='utf-8') as f:
    f.write(resumen_md)
print(f"   Resumen guardado: {ARCHIVO_RESUMEN}")

# ═══════════════════════════════════════════════════════════════════
# CIERRE
# ═══════════════════════════════════════════════════════════════════

print()
print("═" * 70)
print(" CONSOLIDACIÓN COMPLETADA")
print("═" * 70)
print(f"  - Excel:        {ARCHIVO_EXCEL}")
print(f"  - Gráficas:     {DIR_GRAFICAS}/ ({len(list(DIR_GRAFICAS.iterdir()))} archivos)")
print(f"  - Resumen MD:   {ARCHIVO_RESUMEN}")
print()
print(" Conclusiones finales:")
if not ranking_principal.empty:
    top1 = ranking_principal.iloc[0]
    print(f"  → [Principal end-to-end] {top1['chunking']} + {top1['retrieval']}")
    print(f"     score_compuesto = {top1['score_compuesto']:.4f}")
if not ranking_generacion.empty:
    top1_g = ranking_generacion.iloc[0]
    print(f"  → [Generación A/B/C/D]   {top1_g['chunking']} + {top1_g['retrieval']}")
    print(f"     score_generacion = {top1_g['score_generacion']:.4f}")
print()
