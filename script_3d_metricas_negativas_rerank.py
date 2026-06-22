"""
=====================================================================
 SCRIPT 3d: EVALUACIÓN DE NEGATIVAS PARA EL PIPELINE CON RE-RANKING
=====================================================================

Versión espejo del script 3b: evalúa el comportamiento del sistema
ante preguntas Negativas, pero sobre los resultados del pipeline con
re-ranking (script 5).

Mismas métricas (NAR y HR), mismo juez (Gemini 2.5 Flash), mismo
prompt cerrado DECLINACIÓN/SUSTANTIVA. La única diferencia es que las
respuestas evaluadas son las generadas a partir de chunks RE-RANKEADOS.

PREGUNTA QUE RESPONDE: ¿el re-ranker mejora o empeora el
comportamiento de seguridad (declinación) ante preguntas fuera de
alcance? Hay dos hipótesis enfrentadas a contrastar:

    H_A: El rerank mejora la declinación porque elimina chunks
         tangenciales que confundían al LLM y le daban "pistas falsas".

    H_B: El rerank empeora la declinación porque al pulir los top-K,
         el LLM ve contextos más coherentes y se confía, respondiendo
         donde no debería.

Solo procesa las negativas de las 3 combinaciones top-3 evaluadas
con re-ranker, así que son ~120 ejecuciones (40 negativas × 3
combinaciones).

Para comparar antes vs después del re-ranker:
    - antes: metricas_negativas.json (output script 3b)
    - después: metricas_negativas_rerank.json (este script)

Input:
    - resultados_rerank.json (output script 5)

Output:
    - metricas_negativas_rerank.json
    - metricas_negativas_rerank_resumen.xlsx
"""

import os
import json
import time
import warnings
from concurrent.futures import ThreadPoolExecutor, as_completed

warnings.filterwarnings('ignore')

import pandas as pd
from langchain_google_genai import ChatGoogleGenerativeAI

# ═══════════════════════════════════════════════════════════════════
# CONFIGURACIÓN
# ═══════════════════════════════════════════════════════════════════

ARCHIVO_RESULTADOS_PIPELINE = 'resultados_rerank.json'
ARCHIVO_SALIDA_JSON = 'metricas_negativas_rerank.json'
ARCHIVO_SALIDA_EXCEL = 'metricas_negativas_rerank_resumen.xlsx'

GEMINI_API_KEY = "AIzaSyAF4ZZ2jjC1nzvxWOgxXh8IP1YJlF6gzCc"

MODELO_LLM_JUEZ = 'gemini-2.5-flash'
MAX_WORKERS = 8
TAMANO_CHECKPOINT = 50  # guardar cada N evaluaciones

MODO_TEST = False
N_TEST = 5


# ═══════════════════════════════════════════════════════════════════
# PROMPT DEL JUEZ
# ═══════════════════════════════════════════════════════════════════

PROMPT_JUEZ = """Tu tarea es clasificar si una respuesta declina contestar por falta de información en el contexto, o si intenta dar una respuesta sustantiva.

Una respuesta es DECLINACIÓN si:
- Indica explícitamente que no hay información suficiente en el contexto
- Reconoce que no puede responder con los datos disponibles
- Dice que el contexto no menciona el tema preguntado
- Pide información adicional para poder responder

Una respuesta es SUSTANTIVA si:
- Da una respuesta concreta a la pregunta
- Aporta datos clínicos o información específica como respuesta
- Aunque mencione limitaciones, intenta dar una respuesta de fondo

Pregunta: {pregunta}

Respuesta a clasificar: {respuesta}

Responde EXCLUSIVAMENTE con una palabra: DECLINACIÓN o SUSTANTIVA."""


# ═══════════════════════════════════════════════════════════════════
# INICIALIZACIÓN DEL LLM
# ═══════════════════════════════════════════════════════════════════

print("═" * 70)
print(" EVALUACIÓN DE PREGUNTAS NEGATIVAS")
print("═" * 70)

print(" - Inicializando juez (Gemini 2.5 Flash)...")
llm = ChatGoogleGenerativeAI(
    model=MODELO_LLM_JUEZ,
    google_api_key=GEMINI_API_KEY,
    temperature=0.0,
    max_retries=5,
    timeout=120,
)


# ═══════════════════════════════════════════════════════════════════
# JUEZ
# ═══════════════════════════════════════════════════════════════════

def juzgar(pregunta, respuesta):
    """Devuelve 'declinacion' o 'sustantiva' usando el LLM juez."""
    prompt = PROMPT_JUEZ.format(pregunta=pregunta, respuesta=respuesta)
    try:
        msg = llm.invoke(prompt)
        contenido = msg.content.strip().upper()
        if 'DECLINACIÓN' in contenido or 'DECLINACION' in contenido:
            return 'declinacion'
        elif 'SUSTANTIVA' in contenido:
            return 'sustantiva'
        else:
            return 'indeterminado'
    except Exception as e:
        return f'error: {type(e).__name__}'


def procesar_ejecucion(r):
    """Evalúa una ejecución del pipeline (debe ser tipo Negativa)."""
    veredicto = juzgar(r['pregunta'], r['respuesta_generada'])
    return {
        'pregunta_idx': r['pregunta_idx'],
        'tipo': r['tipo'],
        'source': r['source'],
        'chunking': r['chunking'],
        'retrieval': r['retrieval'],
        'veredicto': veredicto,
        'declino_correctamente': veredicto == 'declinacion',
        'alucino': veredicto == 'sustantiva',
    }


# ═══════════════════════════════════════════════════════════════════
# CHECKPOINT
# ═══════════════════════════════════════════════════════════════════

def cargar_checkpoint():
    if not os.path.exists(ARCHIVO_SALIDA_JSON):
        return []
    with open(ARCHIVO_SALIDA_JSON, 'r', encoding='utf-8') as f:
        return json.load(f)


def guardar_checkpoint(metricas):
    with open(ARCHIVO_SALIDA_JSON, 'w', encoding='utf-8') as f:
        json.dump(metricas, f, ensure_ascii=False, indent=2)


def clave(r):
    return f"{r['pregunta_idx']}__{r['chunking']}__{r['retrieval']}"


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════

def main():
    # Cargar resultados pipeline
    with open(ARCHIVO_RESULTADOS_PIPELINE, 'r', encoding='utf-8') as f:
        resultados_pipeline = json.load(f)

    # Filtrar solo Negativas
    negativas = [r for r in resultados_pipeline if r['tipo'] == 'Negativa']
    print(f" - Total resultados pipeline: {len(resultados_pipeline)}")
    print(f" - Ejecuciones Negativas:     {len(negativas)}")

    if MODO_TEST:
        negativas = negativas[:N_TEST]
        print(f"\n   *** MODO TEST: solo {N_TEST} ejecuciones ***\n")

    # Checkpoint
    metricas_previas = cargar_checkpoint()
    hechas = {clave(m) for m in metricas_previas}
    pendientes = [r for r in negativas if clave(r) not in hechas]
    print(f" - Ya evaluadas (checkpoint): {len(hechas)}")
    print(f" - Pendientes:                {len(pendientes)}")
    print()

    if not pendientes:
        print(" - No hay nada nuevo que evaluar.")
        if metricas_previas:
            generar_excel(metricas_previas)
        return

    metricas = list(metricas_previas)
    t_inicio = time.time()
    contador_local = 0

    print(f" - Lanzando {MAX_WORKERS} workers en paralelo...")
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futuros = {executor.submit(procesar_ejecucion, r): r for r in pendientes}
        for futuro in as_completed(futuros):
            r = futuros[futuro]
            try:
                resultado = futuro.result()
                metricas.append(resultado)
                contador_local += 1

                # Mostrar progreso
                if contador_local % 10 == 0 or contador_local == 1:
                    progreso = contador_local / len(pendientes) * 100
                    velocidad = contador_local / max(time.time() - t_inicio, 1) * 60
                    eta = (len(pendientes) - contador_local) / max(velocidad / 60, 0.001) / 60
                    print(f"   [{contador_local}/{len(pendientes)} | "
                          f"{progreso:.1f}% | {velocidad:.0f}/min | "
                          f"ETA {eta:.1f} min]")

                # Checkpoint periódico
                if contador_local % TAMANO_CHECKPOINT == 0:
                    guardar_checkpoint(metricas)
            except Exception as e:
                print(f"   [ERROR] {clave(r)}: {type(e).__name__}: {e}")

    guardar_checkpoint(metricas)
    t_total = (time.time() - t_inicio) / 60
    print(f"\n - Tiempo transcurrido: {t_total:.1f} min")
    print(f" - Total evaluadas: {len(metricas)}")
    print(f" - Archivo: {ARCHIVO_SALIDA_JSON}")
    print()

    generar_excel(metricas)


# ═══════════════════════════════════════════════════════════════════
# AGREGACIÓN Y EXCEL
# ═══════════════════════════════════════════════════════════════════

def generar_excel(metricas):
    df = pd.DataFrame(metricas)

    def agregar(df_in, agrupar_por):
        agg = (df_in
               .groupby(agrupar_por)
               .agg(
                   n_negativas=('declino_correctamente', 'count'),
                   declinacion_rate=('declino_correctamente', 'mean'),
                   hallucination_rate=('alucino', 'mean'),
               )
               .round(4)
               .reset_index())
        return agg

    df_global = agregar(df, ['chunking', 'retrieval'])
    df_por_chunking = agregar(df, ['chunking'])
    df_por_retrieval = agregar(df, ['retrieval'])

    print("═" * 70)
    print(" RESUMEN GLOBAL (chunking × retrieval)")
    print("═" * 70)
    print(df_global.to_string(index=False))
    print()
    print("═" * 70)
    print(" AGREGADO POR CHUNKING")
    print("═" * 70)
    print(df_por_chunking.to_string(index=False))
    print()
    print("═" * 70)
    print(" AGREGADO POR RETRIEVAL")
    print("═" * 70)
    print(df_por_retrieval.to_string(index=False))

    # Excel
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    thin = Side(style='thin', color='BBBBBB')
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    color_titulo = '1F3864'
    color_header = 'CCCCCC'

    def estilo(cell, bold=False, bg=None, color='000000'):
        cell.font = Font(name='Arial', bold=bold, color=color, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = borde
        if bg:
            cell.fill = PatternFill('solid', start_color=bg)

    def añadir_hoja(nombre, df_h, titulo):
        ws = wb.create_sheet(nombre)
        n_cols = len(df_h.columns)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        c = ws.cell(row=1, column=1, value=titulo)
        estilo(c, bold=True, bg=color_titulo, color='FFFFFF')
        ws.row_dimensions[1].height = 24
        for col, cab in enumerate(df_h.columns, 1):
            estilo(ws.cell(row=2, column=col, value=str(cab)),
                   bold=True, bg=color_header)
            ws.column_dimensions[get_column_letter(col)].width = 18
        for fila, (_, row) in enumerate(df_h.iterrows(), 3):
            for col, val in enumerate(row, 1):
                if isinstance(val, float):
                    val = round(val, 4) if pd.notna(val) else '-'
                estilo(ws.cell(row=fila, column=col, value=val))

    añadir_hoja('Global', df_global, 'NEGATIVAS — VISTA GLOBAL (chunking × retrieval)')
    añadir_hoja('Por Chunking', df_por_chunking, 'NEGATIVAS — POR CHUNKING')
    añadir_hoja('Por Retrieval', df_por_retrieval, 'NEGATIVAS — POR RETRIEVAL')

    wb.save(ARCHIVO_SALIDA_EXCEL)
    print(f"\n - Excel guardado: {ARCHIVO_SALIDA_EXCEL}")


if __name__ == '__main__':
    main()
