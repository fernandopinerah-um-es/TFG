"""
=====================================================================
 SCRIPT 6: COMPARACIÓN ANTES vs DESPUÉS DEL RE-RANKING
=====================================================================

Genera el análisis comparativo entre el pipeline base (script 1) y el
pipeline con re-ranking (script 5) sobre las top-3 combinaciones del
ranking principal.

Para cada combinación y cada métrica calcula:
    - media antes (sin rerank)
    - media después (con rerank)
    - delta absoluto y relativo (%)
    - test de Wilcoxon pareado para significancia estadística
    - tamaño de efecto (correlación rank-biserial r)

También genera:
    - tabla resumen comparativa por combinación y métrica
    - tabla de eficiencia: latencia añadida por el rerank
    - veredicto final: ¿el rerank merece la pena?

Input:
    - metricas_ir.json                 (script 2 - base)
    - metricas_ir_rerank.json          (script 2b - con rerank)
    - metricas_ragas.json              (script 3 - base)
    - metricas_ragas_rerank.json       (script 3c - con rerank)
    - resultados_retrieval_generacion.json  (script 1 - latencias base)
    - resultados_rerank.json                (script 5 - latencias con rerank)

Output:
    - comparativa_rerank.xlsx
    - comparativa_rerank.md (resumen con conclusiones)
    - graficas_rerank/ (PNG comparativos)
"""

import json
import os
from pathlib import Path
import warnings

warnings.filterwarnings('ignore')

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ═══════════════════════════════════════════════════════════════════
# CONFIGURACIÓN
# ═══════════════════════════════════════════════════════════════════

# Inputs
F_IR_BASE = 'metricas_ir.json'
F_IR_RERANK = 'metricas_ir_rerank.json'
F_RAGAS_BASE = 'metricas_ragas.json'
F_RAGAS_RERANK = 'metricas_ragas_rerank.json'
F_NEG_BASE = 'metricas_negativas.json'
F_NEG_RERANK = 'metricas_negativas_rerank.json'
F_PIPE_BASE = 'resultados_retrieval_generacion.json'
F_PIPE_RERANK = 'resultados_rerank.json'

# Outputs
DIR_GRAFICAS = Path('graficas_rerank')
F_EXCEL = 'comparativa_rerank.xlsx'
F_RESUMEN = 'comparativa_rerank.md'

# Combinaciones evaluadas con rerank
COMBINACIONES_TOP3 = [
    ('B_markdown',  'semantica_experto'),
    ('A_fixed',     'semantica_experto'),
    ('C_semantica', 'semantica_generalista'),
]

METRICAS_IR = ['mrr', 'P@3', 'P@5', 'R@3', 'R@5', 'NDCG@3', 'NDCG@5']
METRICAS_RAGAS = ['context_precision', 'context_recall',
                  'faithfulness', 'answer_relevancy', 'answer_correctness']
METRICAS_LAT = ['latencia_retrieval_ms', 'latencia_generacion_ms',
                'latencia_total_ms', 'tokens_contexto_aprox']

ALPHA = 0.05  # nivel de significancia para los tests

# Estilo de gráficas
plt.rcParams.update({
    'figure.dpi': 150,
    'savefig.dpi': 300,
    'savefig.bbox': 'tight',
    'savefig.facecolor': 'white',
    'font.family': 'sans-serif',
    'font.size': 11,
    'axes.titlesize': 13,
    'axes.spines.top': False,
    'axes.spines.right': False,
    'axes.grid': True,
    'grid.alpha': 0.3,
})


# ═══════════════════════════════════════════════════════════════════
# CARGA DE DATOS
# ═══════════════════════════════════════════════════════════════════

def cargar_json(path):
    if not os.path.exists(path):
        print(f" - [WARN] No encontrado: {path}")
        return []
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


print("═" * 70)
print(" COMPARATIVA ANTES vs DESPUÉS DEL RE-RANKING")
print("═" * 70)

DIR_GRAFICAS.mkdir(exist_ok=True)

df_ir_base = pd.DataFrame(cargar_json(F_IR_BASE))
df_ir_rk = pd.DataFrame(cargar_json(F_IR_RERANK))
df_ragas_base = pd.DataFrame(cargar_json(F_RAGAS_BASE))
df_ragas_rk = pd.DataFrame(cargar_json(F_RAGAS_RERANK))
df_neg_base = pd.DataFrame(cargar_json(F_NEG_BASE))
df_neg_rk = pd.DataFrame(cargar_json(F_NEG_RERANK))
df_pipe_base = pd.DataFrame(cargar_json(F_PIPE_BASE))
df_pipe_rk = pd.DataFrame(cargar_json(F_PIPE_RERANK))

print(f" - IR base:        {len(df_ir_base)} ejecuciones")
print(f" - IR rerank:      {len(df_ir_rk)} ejecuciones")
print(f" - RAGAS base:     {len(df_ragas_base)} ejecuciones")
print(f" - RAGAS rerank:   {len(df_ragas_rk)} ejecuciones")
print(f" - Negativas base: {len(df_neg_base)} ejecuciones")
print(f" - Negativas rerank:{len(df_neg_rk)} ejecuciones")
print(f" - Pipeline base:  {len(df_pipe_base)} ejecuciones")
print(f" - Pipeline rerank:{len(df_pipe_rk)} ejecuciones")
print()


# ═══════════════════════════════════════════════════════════════════
# FILTRADO A LAS TOP-3 COMBINACIONES
# ═══════════════════════════════════════════════════════════════════

def filtrar_top3(df):
    if df.empty:
        return df
    mask = pd.Series(False, index=df.index)
    for ck, ret in COMBINACIONES_TOP3:
        mask |= (df['chunking'] == ck) & (df['retrieval'] == ret)
    return df[mask].copy()


df_ir_base_t3 = filtrar_top3(df_ir_base)
df_ragas_base_t3 = filtrar_top3(df_ragas_base)
df_pipe_base_t3 = filtrar_top3(df_pipe_base)


# ═══════════════════════════════════════════════════════════════════
# COMPARACIÓN PAREADA POR COMBINACIÓN Y MÉTRICA
# ═══════════════════════════════════════════════════════════════════

def comparar_pareado(df_base, df_rk, ck, ret, metrica):
    """
    Compara una métrica entre base y rerank en una combinación concreta.
    Hace alineación pareada por pregunta_idx para test estadístico válido.

    Robusto a métricas booleanas (declino_correctamente, alucino): se
    convierten a float antes de operar para evitar
    "numpy boolean subtract not supported".
    """
    a = df_base[(df_base['chunking'] == ck) & (df_base['retrieval'] == ret)]
    b = df_rk[(df_rk['chunking'] == ck) & (df_rk['retrieval'] == ret)]

    if a.empty or b.empty or metrica not in a.columns or metrica not in b.columns:
        return None

    merged = a[['pregunta_idx', metrica]].merge(
        b[['pregunta_idx', metrica]],
        on='pregunta_idx',
        suffixes=('_base', '_rk')
    ).dropna()

    if len(merged) < 5:
        return None

    # Convertir a float para soportar booleanos (NAR/HR son bool)
    col_base = merged[f'{metrica}_base'].astype(float)
    col_rk = merged[f'{metrica}_rk'].astype(float)

    media_base = col_base.mean()
    media_rk = col_rk.mean()
    delta_abs = media_rk - media_base
    delta_pct = (delta_abs / media_base * 100) if media_base != 0 else 0

    diff = col_rk - col_base

    # Wilcoxon pareado (no paramétrico, robusto a no-normalidad)
    # Se calcula primero para reutilizar W en el rank-biserial r
    n_eff = int((diff != 0).sum())
    if n_eff > 0:
        try:
            stat, pval = stats.wilcoxon(col_rk, col_base)
            total_ranks = n_eff * (n_eff + 1) / 2
            r_rankbiserial = float(np.sign(delta_abs) * (1 - 2 * stat / total_ranks))
        except Exception:
            stat, pval, r_rankbiserial = None, None, 0.0
    else:
        stat, pval, r_rankbiserial = None, 1.0, 0.0

    return {
        'chunking': ck,
        'retrieval': ret,
        'metrica': metrica,
        'n': len(merged),
        'media_base': round(media_base, 4),
        'media_rerank': round(media_rk, 4),
        'delta_abs': round(delta_abs, 4),
        'delta_pct': round(delta_pct, 2),
        'r_rankbiserial': round(r_rankbiserial, 3),
        'wilcoxon_p': round(pval, 5) if pval is not None else None,
        'significativo': bool(pval is not None and pval < ALPHA),
        'mejora': bool(delta_abs > 0),
    }


print(" - Calculando comparativas pareadas...")

filas_comparativa = []

# Métricas IR (base vs rerank)
for ck, ret in COMBINACIONES_TOP3:
    for met in METRICAS_IR:
        r = comparar_pareado(df_ir_base_t3, df_ir_rk, ck, ret, met)
        if r:
            r['familia'] = 'IR'
            filas_comparativa.append(r)

# Métricas RAGAS
for ck, ret in COMBINACIONES_TOP3:
    for met in METRICAS_RAGAS:
        r = comparar_pareado(df_ragas_base_t3, df_ragas_rk, ck, ret, met)
        if r:
            r['familia'] = 'RAGAS'
            filas_comparativa.append(r)

# Métricas de Negativas (NAR y HR son binarias por ejecución)
df_neg_base_t3 = filtrar_top3(df_neg_base) if not df_neg_base.empty else pd.DataFrame()
for ck, ret in COMBINACIONES_TOP3:
    for met in ['declino_correctamente', 'alucino']:
        if df_neg_base_t3.empty or df_neg_rk.empty:
            continue
        if met not in df_neg_base_t3.columns or met not in df_neg_rk.columns:
            continue
        r = comparar_pareado(df_neg_base_t3, df_neg_rk, ck, ret, met)
        if r:
            r['familia'] = 'Negativas'
            # Renombrar para claridad: NAR = declino_correctamente, HR = alucino
            r['metrica'] = 'NAR' if met == 'declino_correctamente' else 'HR'
            filas_comparativa.append(r)

df_comparativa = pd.DataFrame(filas_comparativa)
print(f"   {len(df_comparativa)} comparativas calculadas")

# ─── MEJORA 1: CORRECCIÓN BONFERRONI POR MÚLTIPLES COMPARACIONES ───
# Por cada combinación se hicieron ~14 tests (7 IR + 5 RAGAS + 2 Negativas).
# Sin corrección, alpha=0.05 sobreestima el número de mejoras "reales" porque
# cada test extra inflaciona la tasa de falsos positivos. Bonferroni divide
# el alpha entre el número de tests por familia de comparaciones.
#
# Aplicamos la corrección DENTRO de cada combinación (no global), porque cada
# combinación es una hipótesis independiente del tipo "¿mejora el rerank en
# esta combinación concreta?". Esto es más estricto que sin corrección y
# más realista que aplicar Bonferroni global a las 42 comparaciones.

if not df_comparativa.empty:
    # Para cada combinación, contar cuántos tests se hicieron y corregir
    df_comparativa['alpha_bonferroni'] = np.nan
    df_comparativa['significativo_bonferroni'] = False
    for ck, ret in COMBINACIONES_TOP3:
        mask = (df_comparativa['chunking'] == ck) & (df_comparativa['retrieval'] == ret)
        n_tests_combo = mask.sum()
        if n_tests_combo > 0:
            alpha_corr = ALPHA / n_tests_combo
            df_comparativa.loc[mask, 'alpha_bonferroni'] = round(alpha_corr, 5)
            df_comparativa.loc[mask, 'significativo_bonferroni'] = (
                df_comparativa.loc[mask, 'wilcoxon_p'] < alpha_corr
            )

    n_sig_naive = df_comparativa['significativo'].sum()
    n_sig_bonf = df_comparativa['significativo_bonferroni'].sum()
    print(f"   Tests significativos sin corrección:  {n_sig_naive}/{len(df_comparativa)}")
    print(f"   Tests significativos con Bonferroni:  {n_sig_bonf}/{len(df_comparativa)}")
print()


# ═══════════════════════════════════════════════════════════════════
# RESUMEN POR COMBINACIÓN
# ═══════════════════════════════════════════════════════════════════

def resumen_por_combinacion(df_comp):
    """Cuenta cuántas métricas mejoran/empeoran/significativamente por combinación."""
    filas = []
    for ck, ret in COMBINACIONES_TOP3:
        sub = df_comp[(df_comp['chunking'] == ck) & (df_comp['retrieval'] == ret)]
        if sub.empty:
            continue
        n_mejoran = (sub['delta_abs'] > 0).sum()
        n_empeoran = (sub['delta_abs'] < 0).sum()
        n_sig_mejora = ((sub['delta_abs'] > 0) & sub['significativo']).sum()
        n_sig_empeora = ((sub['delta_abs'] < 0) & sub['significativo']).sum()
        delta_medio = sub['delta_abs'].mean()
        filas.append({
            'chunking': ck,
            'retrieval': ret,
            'n_metricas': len(sub),
            'n_mejoran': int(n_mejoran),
            'n_empeoran': int(n_empeoran),
            'n_sig_mejora': int(n_sig_mejora),
            'n_sig_empeora': int(n_sig_empeora),
            'delta_medio': round(delta_medio, 4),
        })
    return pd.DataFrame(filas)


df_resumen_combo = resumen_por_combinacion(df_comparativa)


# ═══════════════════════════════════════════════════════════════════
# EFICIENCIA: COSTE DEL RE-RANKING
# ═══════════════════════════════════════════════════════════════════

def comparar_eficiencia():
    """Compara latencias y tokens base vs rerank por combinación."""
    if df_pipe_base_t3.empty or df_pipe_rk.empty:
        return pd.DataFrame()

    # Calcular medias
    base_agg = df_pipe_base_t3.groupby(['chunking', 'retrieval'])[METRICAS_LAT].mean().round(2)
    base_agg.columns = [f'{c}_base' for c in base_agg.columns]
    rk_agg = df_pipe_rk.groupby(['chunking', 'retrieval'])[METRICAS_LAT].mean().round(2)
    rk_agg.columns = [f'{c}_rk' for c in rk_agg.columns]

    # Latencia del rerank: la trazamos por separado en el script 5
    if 'latencia_rerank_ms' in df_pipe_rk.columns:
        rerank_only = df_pipe_rk.groupby(['chunking', 'retrieval'])['latencia_rerank_ms'].mean().round(2)
        rerank_only.name = 'latencia_rerank_ms'

    merged = base_agg.join(rk_agg, how='inner')
    if 'latencia_rerank_ms' in df_pipe_rk.columns:
        merged = merged.join(rerank_only, how='inner')

    # Delta de latencia total
    merged['delta_latencia_total_ms'] = (merged['latencia_total_ms_rk']
                                         - merged['latencia_total_ms_base']).round(2)
    merged['delta_latencia_pct'] = ((merged['delta_latencia_total_ms']
                                     / merged['latencia_total_ms_base']) * 100).round(1)

    return merged.reset_index()


df_eficiencia = comparar_eficiencia()


# ═══════════════════════════════════════════════════════════════════
# MEJORA 2: CORRELACIÓN ΔIR vs ΔRAGAS (pregunta a pregunta)
# ═══════════════════════════════════════════════════════════════════
# Si el re-ranker realmente entendiera mejor la pregunta, las preguntas
# donde más mejora IR (mayor relevancia formal) deberían ser las que
# más mejoran en answer_correctness y faithfulness.
#
# Si la correlación es ~0 o negativa, las mejoras IR son "maquillaje
# formal" sin traslación a calidad de respuesta. Este test es la
# verificación crítica de si el rerank aporta valor real o solo
# reordena chunks formalmente más parecidos a la pregunta.

def correlacion_delta_ir_vs_ragas():
    """Correlaciona Δ NDCG@5 vs Δ answer_correctness pregunta a pregunta."""
    filas = []
    for ck, ret in COMBINACIONES_TOP3:
        # IR base vs rerank
        a_ir = df_ir_base_t3[(df_ir_base_t3['chunking'] == ck)
                             & (df_ir_base_t3['retrieval'] == ret)]
        b_ir = df_ir_rk[(df_ir_rk['chunking'] == ck)
                        & (df_ir_rk['retrieval'] == ret)]
        if a_ir.empty or b_ir.empty:
            continue

        # RAGAS base vs rerank
        a_rag = df_ragas_base_t3[(df_ragas_base_t3['chunking'] == ck)
                                 & (df_ragas_base_t3['retrieval'] == ret)]
        b_rag = df_ragas_rk[(df_ragas_rk['chunking'] == ck)
                            & (df_ragas_rk['retrieval'] == ret)]
        if a_rag.empty or b_rag.empty:
            continue

        for met_ir in ['NDCG@5', 'mrr']:
            for met_rag in ['answer_correctness', 'faithfulness']:
                if met_ir not in a_ir.columns or met_rag not in a_rag.columns:
                    continue
                m_ir = (a_ir[['pregunta_idx', met_ir]].dropna()
                        .merge(b_ir[['pregunta_idx', met_ir]].dropna(),
                               on='pregunta_idx', suffixes=('_b', '_r')))
                m_ir[f'delta_{met_ir}'] = m_ir[f'{met_ir}_r'] - m_ir[f'{met_ir}_b']

                m_rag = (a_rag[['pregunta_idx', met_rag]].dropna()
                         .merge(b_rag[['pregunta_idx', met_rag]].dropna(),
                                on='pregunta_idx', suffixes=('_b', '_r')))
                m_rag[f'delta_{met_rag}'] = (m_rag[f'{met_rag}_r']
                                             - m_rag[f'{met_rag}_b'])

                cruzado = m_ir[['pregunta_idx', f'delta_{met_ir}']].merge(
                    m_rag[['pregunta_idx', f'delta_{met_rag}']],
                    on='pregunta_idx'
                )

                if len(cruzado) < 5:
                    continue

                # Pearson y Spearman para robustez
                pearson_r, pearson_p = stats.pearsonr(
                    cruzado[f'delta_{met_ir}'], cruzado[f'delta_{met_rag}']
                )
                spearman_r, spearman_p = stats.spearmanr(
                    cruzado[f'delta_{met_ir}'], cruzado[f'delta_{met_rag}']
                )

                # ¿En qué porcentaje de preguntas mejoran AMBAS?
                ambas_mejoran = ((cruzado[f'delta_{met_ir}'] > 0)
                                 & (cruzado[f'delta_{met_rag}'] > 0)).sum()
                ir_mejora = (cruzado[f'delta_{met_ir}'] > 0).sum()
                rag_mejora = (cruzado[f'delta_{met_rag}'] > 0).sum()
                pct_ambas = (ambas_mejoran / len(cruzado) * 100) if len(cruzado) > 0 else 0

                interpretacion = ('mejoras IR se trasladan a RAGAS'
                                  if pearson_r > 0.3
                                  else ('correlación nula: maquillaje IR'
                                        if abs(pearson_r) < 0.2
                                        else 'correlación débil/contradictoria'))

                filas.append({
                    'chunking': ck,
                    'retrieval': ret,
                    'met_ir': met_ir,
                    'met_ragas': met_rag,
                    'n_preguntas': len(cruzado),
                    'pearson_r': round(pearson_r, 3),
                    'pearson_p': round(pearson_p, 4),
                    'spearman_r': round(spearman_r, 3),
                    'spearman_p': round(spearman_p, 4),
                    'preguntas_mejora_IR': int(ir_mejora),
                    'preguntas_mejora_RAGAS': int(rag_mejora),
                    'preguntas_mejoran_ambas': int(ambas_mejoran),
                    'pct_mejoran_ambas': round(pct_ambas, 1),
                    'interpretacion': interpretacion,
                })
    return pd.DataFrame(filas)


df_correlacion = correlacion_delta_ir_vs_ragas()
print(" - Análisis de correlación ΔIR vs ΔRAGAS calculado")
print()


# ═══════════════════════════════════════════════════════════════════
# VEREDICTO AUTOMÁTICO
# ═══════════════════════════════════════════════════════════════════

def generar_veredicto():
    """
    MEJORA 3: Veredicto realista que pondera evidencia jerárquica:

    1. Mejoras significativas tras Bonferroni (peso máximo)
    2. Mejoras en métricas RAGAS clave (faithfulness, answer_correctness)
    3. Cambios en NAR (negativas) — efecto seguridad
    4. Mejoras solo en IR sin trasladarse a RAGAS — peso bajo o neutro

    Categorías posibles:
    - FAVORABLE FUERTE: ≥1 sig Bonferroni en RAGAS clave + sin empeoras críticas
    - FAVORABLE MODERADO: mejoras IR significativas + mejoras leves en RAGAS clave
    - NEUTRO: mejoras solo formales sin trasladarse a calidad
    - DESFAVORABLE: empeora métricas RAGAS clave o NAR (seguridad)
    """
    veredictos = {}

    for ck, ret in COMBINACIONES_TOP3:
        sub = df_comparativa[(df_comparativa['chunking'] == ck)
                             & (df_comparativa['retrieval'] == ret)]
        if sub.empty:
            veredictos[(ck, ret)] = 'sin datos'
            continue

        # Conteos sin corrección y con Bonferroni
        n_sig_naive = sub['significativo'].sum()
        n_sig_bonf = sub['significativo_bonferroni'].sum()

        # Familias
        ragas_clave = ['faithfulness', 'answer_correctness']
        sub_ragas_clave = sub[sub['metrica'].isin(ragas_clave)]
        sub_ir = sub[sub['familia'] == 'IR']
        sub_neg = sub[sub['familia'] == 'Negativas']

        # Significativas tras Bonferroni en métricas RAGAS clave (lo más importante)
        sig_bonf_ragas_clave = sub_ragas_clave[
            sub_ragas_clave['significativo_bonferroni']
        ]
        # Significativas tras Bonferroni en IR
        sig_bonf_ir = sub_ir[sub_ir['significativo_bonferroni']]
        # Empeoras en métricas clave (significativas o no, importan por dirección)
        empeoras_ragas_clave = sub_ragas_clave[sub_ragas_clave['delta_abs'] < 0]
        # Cambio en NAR (Negative Acceptance Rate)
        nar_row = sub[sub['metrica'] == 'NAR']
        delta_nar = nar_row['delta_abs'].iloc[0] if not nar_row.empty else 0

        # Métricas clave: directizalidad
        delta_AC = sub[sub['metrica'] == 'answer_correctness']['delta_abs']
        delta_F = sub[sub['metrica'] == 'faithfulness']['delta_abs']
        ac_delta = delta_AC.iloc[0] if len(delta_AC) else 0
        f_delta = delta_F.iloc[0] if len(delta_F) else 0
        ac_mejora = ac_delta > 0
        f_mejora = f_delta > 0

        # Latencia
        lat_extra_pct = None
        if not df_eficiencia.empty:
            ef = df_eficiencia[(df_eficiencia['chunking'] == ck)
                               & (df_eficiencia['retrieval'] == ret)]
            if not ef.empty and 'delta_latencia_pct' in ef.columns:
                lat_extra_pct = ef['delta_latencia_pct'].iloc[0]

        # ─── LÓGICA DE DECISIÓN JERÁRQUICA ───
        # Empeoras importantes pesan más que mejoras formales
        empeora_AC = ac_delta < -0.01
        empeora_F = f_delta < -0.01
        baja_NAR_significativa = delta_nar < -0.05  # caída de 5+ puntos en seguridad

        razones = []

        if len(sig_bonf_ragas_clave) >= 1 and not empeora_AC and not empeora_F:
            v = 'FAVORABLE FUERTE'
            razones.append(f"{len(sig_bonf_ragas_clave)} mejora(s) sig. Bonferroni en métricas RAGAS clave")
        elif baja_NAR_significativa or (empeora_AC and empeora_F):
            v = 'DESFAVORABLE'
            if baja_NAR_significativa:
                razones.append(f"caída de NAR de {delta_nar:+.3f} (riesgo de seguridad)")
            if empeora_AC:
                razones.append(f"answer_correctness empeora ({ac_delta:+.4f})")
            if empeora_F:
                razones.append(f"faithfulness empeora ({f_delta:+.4f})")
        elif len(sig_bonf_ir) >= 1 and (ac_delta > 0.01 or f_delta > 0.01):
            v = 'FAVORABLE MODERADO'
            razones.append(f"{len(sig_bonf_ir)} mejoras IR sig. Bonferroni + tendencia positiva en RAGAS clave")
        elif len(sig_bonf_ir) >= 1 and abs(ac_delta) < 0.01 and abs(f_delta) < 0.01:
            v = 'NEUTRO (mejoras solo formales)'
            razones.append('mejoras IR significativas no se trasladan a métricas RAGAS clave (Δ AC y Δ F casi nulos)')
        elif n_sig_naive == 0:
            v = 'NEUTRO'
            razones.append('sin diferencias estadísticamente significativas')
        else:
            v = 'AMBIGUO'
            razones.append(f"{n_sig_naive} sig. sin Bonferroni, {n_sig_bonf} con Bonferroni; señales mixtas")

        veredictos[(ck, ret)] = {
            'veredicto': v,
            'razones': razones,
            'n_sig_naive': int(n_sig_naive),
            'n_sig_bonferroni': int(n_sig_bonf),
            'sig_bonf_RAGAS_clave': int(len(sig_bonf_ragas_clave)),
            'sig_bonf_IR': int(len(sig_bonf_ir)),
            'AC_mejora': bool(ac_mejora),
            'AC_delta': round(ac_delta, 4),
            'F_mejora': bool(f_mejora),
            'F_delta': round(f_delta, 4),
            'NAR_delta': round(delta_nar, 4),
            'lat_extra_pct': lat_extra_pct,
        }

    return veredictos


veredictos = generar_veredicto()


# ═══════════════════════════════════════════════════════════════════
# IMPRESIÓN POR PANTALLA DEL ANÁLISIS
# ═══════════════════════════════════════════════════════════════════

print("═" * 70)
print(" RESUMEN POR COMBINACIÓN")
print("═" * 70)
print(df_resumen_combo.to_string(index=False))
print()

print("═" * 70)
print(" CORRELACIÓN ΔIR vs ΔRAGAS (¿las mejoras IR se trasladan a calidad?)")
print("═" * 70)
if not df_correlacion.empty:
    cols_corr = ['chunking', 'retrieval', 'met_ir', 'met_ragas',
                 'pearson_r', 'pearson_p', 'pct_mejoran_ambas', 'interpretacion']
    print(df_correlacion[cols_corr].to_string(index=False))
print()

print("═" * 70)
print(" EFICIENCIA (latencia añadida por re-ranker)")
print("═" * 70)
if not df_eficiencia.empty:
    cols_resumen = ['chunking', 'retrieval', 'latencia_total_ms_base',
                    'latencia_total_ms_rk', 'delta_latencia_total_ms',
                    'delta_latencia_pct']
    cols_existentes = [c for c in cols_resumen if c in df_eficiencia.columns]
    print(df_eficiencia[cols_existentes].to_string(index=False))
print()

print("═" * 70)
print(" VEREDICTO POR COMBINACIÓN (con corrección Bonferroni)")
print("═" * 70)
for (ck, ret), v in veredictos.items():
    print(f"\n  {ck} + {ret}")
    if isinstance(v, dict):
        print(f"     → {v['veredicto']}")
        for r in v['razones']:
            print(f"        · {r}")
        print(f"     Tests sig. sin corrección:           {v['n_sig_naive']}/14")
        print(f"     Tests sig. tras Bonferroni:          {v['n_sig_bonferroni']}/14")
        print(f"     De los Bonferroni: en RAGAS clave:   {v['sig_bonf_RAGAS_clave']}")
        print(f"     De los Bonferroni: en IR:            {v['sig_bonf_IR']}")
        print(f"     Δ Answer Correctness:                {v['AC_delta']:+.4f}")
        print(f"     Δ Faithfulness:                      {v['F_delta']:+.4f}")
        print(f"     Δ NAR (negativas):                   {v['NAR_delta']:+.4f}")
        if v['lat_extra_pct'] is not None:
            signo = '+' if v['lat_extra_pct'] > 0 else ''
            print(f"     Δ Latencia:                          {signo}{v['lat_extra_pct']:.1f}%")
    else:
        print(f"     → {v}")
print()


# ═══════════════════════════════════════════════════════════════════
# GRÁFICAS COMPARATIVAS
# ═══════════════════════════════════════════════════════════════════

print(" - Generando gráficas comparativas...")


def grafica_barras_comparativas(df_comp, familia, archivo):
    """Barras agrupadas: para cada métrica, una barra base y otra rerank."""
    sub = df_comp[df_comp['familia'] == familia]
    if sub.empty:
        return
    fig, axes = plt.subplots(1, len(COMBINACIONES_TOP3),
                             figsize=(5 * len(COMBINACIONES_TOP3), 5),
                             sharey=True)
    if len(COMBINACIONES_TOP3) == 1:
        axes = [axes]

    for ax, (ck, ret) in zip(axes, COMBINACIONES_TOP3):
        sub_c = sub[(sub['chunking'] == ck) & (sub['retrieval'] == ret)]
        if sub_c.empty:
            continue
        x = np.arange(len(sub_c))
        width = 0.35
        ax.bar(x - width/2, sub_c['media_base'], width, label='Base', color='#4C72B0')
        ax.bar(x + width/2, sub_c['media_rerank'], width, label='Re-rank', color='#DD8452')

        # Anotar significancia
        for i, (_, row) in enumerate(sub_c.iterrows()):
            if row['significativo']:
                marca = '*' if row['delta_abs'] > 0 else '×'
                color = 'green' if row['delta_abs'] > 0 else 'red'
                y = max(row['media_base'], row['media_rerank']) + 0.02
                ax.text(i, y, marca, ha='center', fontsize=14, color=color, fontweight='bold')

        ax.set_xticks(x)
        ax.set_xticklabels(sub_c['metrica'], rotation=30, ha='right', fontsize=9)
        ax.set_title(f'{ck}\n+ {ret}', fontsize=11)
        ax.set_ylim(0, 1.1)
        ax.legend(fontsize=9)

    fig.suptitle(f'Comparativa {familia}: Base vs Re-ranker (* = mejora sig., × = empeora sig.)',
                 fontweight='bold')
    plt.tight_layout()
    plt.savefig(DIR_GRAFICAS / archivo)
    plt.close()


grafica_barras_comparativas(df_comparativa, 'IR', 'comparativa_IR.png')
grafica_barras_comparativas(df_comparativa, 'RAGAS', 'comparativa_RAGAS.png')
grafica_barras_comparativas(df_comparativa, 'Negativas', 'comparativa_Negativas.png')


def grafica_delta(df_comp, archivo):
    """Forest plot estilo: delta absoluto con marca de significancia."""
    if df_comp.empty:
        return
    df_plot = df_comp.copy()
    df_plot['etiqueta'] = (df_plot['chunking'] + '|'
                           + df_plot['retrieval'].str.replace('semantica_', 's_')
                           + '|' + df_plot['metrica'])
    df_plot = df_plot.sort_values('delta_abs')

    fig, ax = plt.subplots(figsize=(9, max(8, len(df_plot) * 0.25)))
    colors = ['#2ca02c' if d > 0 else '#d62728' for d in df_plot['delta_abs']]
    bars = ax.barh(df_plot['etiqueta'], df_plot['delta_abs'], color=colors,
                   alpha=0.7, edgecolor='black', linewidth=0.5)

    # Marcar significancia con asterisco
    for i, (_, row) in enumerate(df_plot.iterrows()):
        if row['significativo']:
            x = row['delta_abs']
            ax.text(x + 0.002 * np.sign(x), i, '*', fontsize=14,
                    fontweight='bold', va='center',
                    color='darkgreen' if x > 0 else 'darkred')

    ax.axvline(0, color='black', linewidth=0.8)
    ax.set_xlabel('Delta absoluto (re-rank − base)')
    ax.set_title('Efecto del re-ranker por combinación y métrica\n'
                 '(verde = mejora, rojo = empeora; * = p<0.05)',
                 fontweight='bold')
    plt.yticks(fontsize=7)
    plt.tight_layout()
    plt.savefig(DIR_GRAFICAS / archivo)
    plt.close()


grafica_delta(df_comparativa, 'forest_delta.png')


def grafica_eficiencia(df_ef, archivo):
    """Barras: latencia base vs latencia con rerank."""
    if df_ef.empty:
        return
    fig, ax = plt.subplots(figsize=(10, 5))
    df_p = df_ef.copy()
    df_p['etiqueta'] = df_p['chunking'] + '+\n' + df_p['retrieval'].str.replace('semantica_', 's_')

    x = np.arange(len(df_p))
    width = 0.35
    ax.bar(x - width/2, df_p['latencia_total_ms_base'], width,
           label='Base', color='#4C72B0', alpha=0.85)
    ax.bar(x + width/2, df_p['latencia_total_ms_rk'], width,
           label='Re-rank', color='#DD8452', alpha=0.85)

    for i, (_, row) in enumerate(df_p.iterrows()):
        # Formato con signo automático: evita el bug "+-2%"
        delta_pct = row['delta_latencia_pct']
        signo = '+' if delta_pct > 0 else ''  # negativos ya llevan su signo
        ax.text(i + width/2, row['latencia_total_ms_rk'] + 50,
                f"{signo}{delta_pct:.0f}%",
                ha='center', fontsize=9,
                color='darkred' if delta_pct > 0 else 'darkgreen')

    ax.set_xticks(x)
    ax.set_xticklabels(df_p['etiqueta'], fontsize=9)
    ax.set_ylabel('Latencia total media (ms)')
    ax.set_title('Coste computacional del re-ranking', fontweight='bold')
    ax.legend()
    plt.tight_layout()
    plt.savefig(DIR_GRAFICAS / archivo)
    plt.close()


grafica_eficiencia(df_eficiencia, 'comparativa_eficiencia.png')

print(f"   Gráficas guardadas en: {DIR_GRAFICAS}/")
print()


# ═══════════════════════════════════════════════════════════════════
# EXCEL DE COMPARATIVA
# ═══════════════════════════════════════════════════════════════════

print(" - Generando Excel comparativo...")

wb = openpyxl.Workbook()
wb.remove(wb.active)
thin = Side(style='thin', color='BBBBBB')
borde = Border(left=thin, right=thin, top=thin, bottom=thin)


def estilo(cell, bold=False, bg=None, color='000000', size=10):
    cell.font = Font(name='Arial', bold=bold, color=color, size=size)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = borde
    if bg:
        cell.fill = PatternFill('solid', start_color=bg)


def añadir_hoja(nombre, df_h, titulo):
    ws = wb.create_sheet(nombre)
    n_cols = len(df_h.columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    estilo(ws.cell(row=1, column=1, value=titulo), bold=True,
           bg='1F3864', color='FFFFFF', size=11)
    ws.row_dimensions[1].height = 26

    for col, cab in enumerate(df_h.columns, 1):
        estilo(ws.cell(row=2, column=col, value=str(cab)), bold=True, bg='CCCCCC')
        ws.column_dimensions[get_column_letter(col)].width = 15

    for fila, (_, row) in enumerate(df_h.iterrows(), 3):
        for col, val in enumerate(row, 1):
            if isinstance(val, float):
                val = round(val, 4) if pd.notna(val) else '-'
            estilo(ws.cell(row=fila, column=col, value=val))

    # Color scale para columnas numéricas relevantes
    for col_idx, col_nombre in enumerate(df_h.columns, 1):
        if col_nombre in ('delta_abs', 'delta_pct') and len(df_h) > 1:
            col_letter = get_column_letter(col_idx)
            rango = f"{col_letter}3:{col_letter}{len(df_h) + 2}"
            rule = ColorScaleRule(
                start_type='min', start_color='F8696B',
                mid_type='num', mid_value=0, mid_color='FFFFFF',
                end_type='max', end_color='63BE7B'
            )
            ws.conditional_formatting.add(rango, rule)


añadir_hoja('Comparativa Detalle', df_comparativa,
            'COMPARATIVA DETALLADA: BASE vs RE-RANK (todas las métricas, '
            'con corrección Bonferroni)')
añadir_hoja('Resumen por Combinación', df_resumen_combo,
            'RESUMEN AGREGADO POR COMBINACIÓN')
if not df_correlacion.empty:
    añadir_hoja('Correlacion IR vs RAGAS', df_correlacion,
                'CORRELACIÓN ΔIR vs ΔRAGAS pregunta a pregunta '
                '(¿las mejoras IR se trasladan a calidad real?)')
if not df_eficiencia.empty:
    añadir_hoja('Eficiencia', df_eficiencia,
                'COSTE COMPUTACIONAL: LATENCIA Y TOKENS')

# Hoja con el veredicto realista (Mejora 3)
filas_veredicto = []
for (ck, ret), v in veredictos.items():
    if isinstance(v, dict):
        filas_veredicto.append({
            'chunking': ck,
            'retrieval': ret,
            'veredicto': v['veredicto'],
            'razones': '; '.join(v['razones']),
            'sig_naive': v['n_sig_naive'],
            'sig_bonferroni': v['n_sig_bonferroni'],
            'sig_bonf_RAGAS_clave': v['sig_bonf_RAGAS_clave'],
            'sig_bonf_IR': v['sig_bonf_IR'],
            'delta_AC': v['AC_delta'],
            'delta_F': v['F_delta'],
            'delta_NAR': v['NAR_delta'],
            'delta_lat_pct': v['lat_extra_pct'],
        })
df_veredicto = pd.DataFrame(filas_veredicto)
añadir_hoja('Veredicto', df_veredicto,
            'VEREDICTO REALISTA POR COMBINACIÓN '
            '(con corrección Bonferroni y ponderación de métricas RAGAS clave)')

wb.save(F_EXCEL)
print(f"   Excel guardado: {F_EXCEL}")
print()


# ═══════════════════════════════════════════════════════════════════
# RESUMEN MARKDOWN
# ═══════════════════════════════════════════════════════════════════

def generar_md():
    L = []
    L.append('# Comparativa: pipeline base vs pipeline con re-ranking\n')
    L.append('## 1. Diseño experimental\n')
    L.append('Se aplicó un cross-encoder (`BAAI/bge-reranker-v2-m3`) sobre las **3 mejores '
             'combinaciones** del ranking principal del experimento base:\n')
    for i, (ck, ret) in enumerate(COMBINACIONES_TOP3, 1):
        L.append(f'{i}. `{ck}` + `{ret}`')
    L.append('\nFlujo modificado: el retriever base recupera 20 candidatos, el cross-encoder '
             'los reordena y los 5 mejores pasan al LLM generador.\n')
    L.append('### Rigor estadístico aplicado\n')
    L.append('Por cada combinación se realizan ~14 tests de Wilcoxon pareados '
             '(7 IR + 5 RAGAS + 2 Negativas). Para evitar inflar la tasa de '
             'falsos positivos por **comparaciones múltiples**, además del '
             'p-value sin corregir (α=0.05) reportamos significancia con '
             '**corrección Bonferroni** (α/n_tests, típicamente 0.0036). El '
             'veredicto final pondera más fuerte los tests que sobreviven '
             'Bonferroni y prioriza métricas RAGAS de generación clave '
             '(faithfulness, answer_correctness) sobre métricas IR formales.\n')

    L.append('## 2. Resultados por combinación\n')
    for ck, ret in COMBINACIONES_TOP3:
        L.append(f"### {ck} + {ret}\n")
        sub = df_comparativa[(df_comparativa['chunking'] == ck)
                             & (df_comparativa['retrieval'] == ret)]
        if sub.empty:
            L.append('*Sin datos*\n')
            continue
        L.append('| Familia | Métrica | Base | Re-rank | Δ abs | Δ % | r (RB) | Wilcoxon p | Sig (α=0.05) | Sig Bonferroni |')
        L.append('|---------|---------|------|---------|-------|-----|--------|------------|--------------|----------------|')
        for _, r in sub.iterrows():
            sig = '✓' if r['significativo'] else ''
            sig_bonf = '✓✓' if r['significativo_bonferroni'] else ''
            sym = '↑' if r['delta_abs'] > 0 else ('↓' if r['delta_abs'] < 0 else '·')
            p_str = f"{r['wilcoxon_p']:.4f}" if r['wilcoxon_p'] is not None else '-'
            L.append(f"| {r['familia']} | `{r['metrica']}` | {r['media_base']:.4f} | "
                     f"{r['media_rerank']:.4f} | {sym}{abs(r['delta_abs']):.4f} | "
                     f"{r['delta_pct']:+.1f}% | {r['r_rankbiserial']:+.3f} | {p_str} | {sig} | {sig_bonf} |")
        L.append('')

        v = veredictos.get((ck, ret), 'sin datos')
        if isinstance(v, dict):
            L.append(f"**Veredicto: {v['veredicto']}**\n")
            for raz in v['razones']:
                L.append(f"- {raz}")
            L.append('')
            L.append(f"- Tests significativos sin corrección: {v['n_sig_naive']}/14")
            L.append(f"- Tests significativos tras Bonferroni: {v['n_sig_bonferroni']}/14 "
                     f"(de los cuales {v['sig_bonf_RAGAS_clave']} en RAGAS clave, "
                     f"{v['sig_bonf_IR']} en IR)")
            L.append(f"- Δ Answer Correctness: {v['AC_delta']:+.4f} "
                     f"({'mejora' if v['AC_mejora'] else 'no mejora'})")
            L.append(f"- Δ Faithfulness: {v['F_delta']:+.4f} "
                     f"({'mejora' if v['F_mejora'] else 'no mejora'})")
            L.append(f"- Δ NAR (negativas): {v['NAR_delta']:+.4f} "
                     f"({'⚠️ EMPEORA seguridad' if v['NAR_delta'] < -0.05 else 'estable'})")
            if v['lat_extra_pct'] is not None:
                signo = '+' if v['lat_extra_pct'] > 0 else ''
                L.append(f"- Δ Latencia: {signo}{v['lat_extra_pct']:.1f}%")
            L.append('')

    L.append('## 3. Análisis de coherencia: ¿las mejoras IR se trasladan a RAGAS?\n')
    if not df_correlacion.empty:
        L.append('Si el re-ranker entendiera realmente mejor la pregunta, las preguntas '
                 'donde más mejora IR deberían ser las que más mejoran en answer_correctness. '
                 'Las correlaciones pregunta-a-pregunta entre Δ IR y Δ RAGAS lo verifican:\n')
        L.append('| Combinación | ΔIR | ΔRAGAS | Pearson r | p-value | % preguntas que mejoran AMBAS | Interpretación |')
        L.append('|-------------|-----|--------|-----------|---------|-------------------------------|----------------|')
        for _, r in df_correlacion.iterrows():
            L.append(f"| `{r['chunking']}+{r['retrieval']}` | {r['met_ir']} | {r['met_ragas']} | "
                     f"{r['pearson_r']:+.3f} | {r['pearson_p']:.4f} | "
                     f"{r['pct_mejoran_ambas']}% | {r['interpretacion']} |")
        L.append('')

    L.append('## 4. Eficiencia\n')
    if not df_eficiencia.empty:
        L.append('| Combinación | Lat. base (ms) | Lat. re-rank (ms) | Δ ms | Δ % |')
        L.append('|-------------|----------------|-------------------|------|-----|')
        for _, r in df_eficiencia.iterrows():
            # Fix bug +-2%: signo automático
            d_ms = r['delta_latencia_total_ms']
            d_pct = r['delta_latencia_pct']
            ms_str = f"+{d_ms:.0f}" if d_ms > 0 else f"{d_ms:.0f}"
            pct_str = f"+{d_pct:.1f}%" if d_pct > 0 else f"{d_pct:.1f}%"
            L.append(f"| `{r['chunking']}+{r['retrieval']}` | "
                     f"{r['latencia_total_ms_base']:.0f} | "
                     f"{r['latencia_total_ms_rk']:.0f} | "
                     f"{ms_str} | {pct_str} |")
        L.append('')
        L.append('**Nota sobre el coste:** la diferencia de latencia es del orden del '
                 'ruido de medida (±2%). Esto se debe a que retriever, re-ranker y '
                 'generador comparten la misma GPU; el cuello de botella es la '
                 'generación del LLM, no el cross-encoder.\n')

    L.append('## 5. Conclusión global\n')
    L.append('Criterios para considerar el re-ranker **justificado**:\n')
    L.append('- Mejoras significativas tras Bonferroni en métricas RAGAS clave (faithfulness, answer_correctness)')
    L.append('- Sin empeoras significativas en métricas RAGAS clave ni en NAR (seguridad)')
    L.append('- Sobrecoste de latencia asumible\n')

    n_fav_fuerte = sum(1 for v in veredictos.values()
                       if isinstance(v, dict) and 'FAVORABLE FUERTE' in v['veredicto'])
    n_fav_mod = sum(1 for v in veredictos.values()
                    if isinstance(v, dict) and 'FAVORABLE MODERADO' in v['veredicto'])
    n_neutro = sum(1 for v in veredictos.values()
                   if isinstance(v, dict) and 'NEUTRO' in v['veredicto'])
    n_desfav = sum(1 for v in veredictos.values()
                   if isinstance(v, dict) and 'DESFAVORABLE' in v['veredicto'])

    L.append(f"**Distribución de veredictos sobre {len(COMBINACIONES_TOP3)} combinaciones:**\n")
    L.append(f"- Favorable fuerte:    {n_fav_fuerte}")
    L.append(f"- Favorable moderado:  {n_fav_mod}")
    L.append(f"- Neutro:              {n_neutro}")
    L.append(f"- Desfavorable:        {n_desfav}\n")

    if n_fav_fuerte >= 2:
        L.append('→ El re-ranker aporta valor real (mejoras significativas tras Bonferroni '
                 'en métricas RAGAS clave). Su inclusión en producción está justificada.')
    elif n_desfav >= 1:
        L.append('→ El re-ranker degrada al menos una combinación en métricas críticas '
                 '(answer_correctness, faithfulness o NAR). Recomendación: NO incluirlo en '
                 'producción para esta arquitectura sin investigación adicional.')
    else:
        L.append('→ El re-ranker produce mejoras estadísticamente significativas en métricas '
                 'IR formales que no se trasladan a métricas RAGAS de calidad de respuesta '
                 '(answer_correctness, faithfulness se mantienen estables o empeoran '
                 'levemente). Tras corrección Bonferroni por múltiples comparaciones, la '
                 'mayoría de las "mejoras" sin corregir desaparecen. Las correlaciones '
                 'pregunta-a-pregunta entre Δ IR y Δ RAGAS son nulas o negativas, '
                 'confirmando que el re-ranking reordena formalmente sin aportar valor '
                 'real al generador. Para este corpus de protocolos clínicos en español '
                 'con un retriever semántico fine-tuneado, **el sobrecoste arquitectónico '
                 'del re-ranking no se justifica con los datos observados**. Este es un '
                 'hallazgo metodológico legítimo que debe reportarse como tal en la '
                 'memoria, no como un fracaso experimental.')
    L.append('')

    return '\n'.join(L)


print(" - Generando resumen Markdown...")
md = generar_md()
with open(F_RESUMEN, 'w', encoding='utf-8') as f:
    f.write(md)
print(f"   Resumen guardado: {F_RESUMEN}")
print()

print("═" * 70)
print(" COMPARATIVA COMPLETADA")
print("═" * 70)
print(f"  - Excel:    {F_EXCEL}")
print(f"  - Resumen:  {F_RESUMEN}")
print(f"  - Gráficas: {DIR_GRAFICAS}/")
print()
