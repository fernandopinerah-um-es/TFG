"""
=====================================================================
 SCRIPT 2: MÉTRICAS IR CLÁSICAS
=====================================================================

Calcula las métricas tradicionales de Information Retrieval sobre los
resultados del pipeline RAG:
    - Precision@K
    - Recall@K
    - MRR
    - NDCG@K
A K=3 y K=5.

CRITERIO DE RELEVANCIA:
    Un chunk recuperado se considera RELEVANTE si su texto está
    "contenido" en la macro-sección de origen de la pregunta, según
    fuzzy partial_ratio >= UMBRAL_FUZZY.

DENOMINADOR DE RECALL:
    Para cada pregunta y estrategia de chunking, se calcula
    empíricamente cuántos chunks del documento están contenidos en
    la macro-sección. Ese es el universo de chunks relevantes.

EXCLUSIONES:
    - D_jerarquico: la dirección del fuzzy se rompe (chunk padre > macro)
    - Negativas: no tienen macro-sección de origen

Input:
    - resultados_retrieval_generacion.json (output script 1)
    - dataset_test.json (macro-secciones)
    - Elasticsearch (para escanear chunks por documento)

Output:
    - metricas_ir.json (resultados completos por ejecución)
    - metricas_ir_resumen.xlsx (tablas agregadas)
"""

import json
import os
from collections import defaultdict
from itertools import product

import numpy as np
import pandas as pd
from elasticsearch import Elasticsearch
from thefuzz import fuzz

# ═══════════════════════════════════════════════════════════════════
# CONFIGURACIÓN
# ═══════════════════════════════════════════════════════════════════

ARCHIVO_RESULTADOS = 'resultados_retrieval_generacion.json'
ARCHIVO_MACROS = 'dataset_test.json'
ARCHIVO_SALIDA_JSON = 'metricas_ir.json'
ARCHIVO_SALIDA_EXCEL = 'metricas_ir_resumen.xlsx'
ARCHIVO_CACHE_DENOMINADORES = 'cache_denominadores_recall.json'

ES_HOST = 'http://localhost:9201'
UMBRAL_FUZZY = 90
KS_EVALUACION = [3, 5]

CHUNKINGS_EVALUADOS = ['A_fixed', 'B_markdown', 'C_semantica']  # D excluido
TIPOS_EVALUADOS = ['Directa', 'Caso', 'Compleja']  # Negativas excluidas

INDICES_POR_CHUNKING = {
    'A_fixed': 'indice_estrategia_a_fixed',
    'B_markdown': 'indice_estrategia_b_markdown',
    'C_semantica': 'indice_estrategia_c_semantica',
}

# ═══════════════════════════════════════════════════════════════════
# CARGA DE DATOS
# ═══════════════════════════════════════════════════════════════════

print("═" * 70)
print(" CÁLCULO DE MÉTRICAS IR CLÁSICAS")
print("═" * 70)

# Resultados del pipeline
with open(ARCHIVO_RESULTADOS, 'r', encoding='utf-8') as f:
    resultados = json.load(f)
print(f" - Resultados pipeline cargados: {len(resultados)} ejecuciones")

# Macro-secciones
with open(ARCHIVO_MACROS, 'r', encoding='utf-8') as f:
    macros = json.load(f)
mapa_macros = {m['macro_id']: m['text'] for m in macros}
print(f" - Macro-secciones cargadas: {len(mapa_macros)}")

# Elasticsearch
es = Elasticsearch(ES_HOST, request_timeout=60)
assert es.ping(), 'ERROR: Elasticsearch no disponible'
print(f" - Elasticsearch conectado en {ES_HOST}")
print()


# ═══════════════════════════════════════════════════════════════════
# DENOMINADOR DE RECALL: chunks relevantes en cada (pregunta, chunking)
# ═══════════════════════════════════════════════════════════════════

def es_relevante(texto_chunk, texto_macro):
    """
    Un chunk es relevante si está 'contenido' en la macro-sección.
    Usamos partial_ratio: chunk (aguja) vs macro (pajar).
    """
    return fuzz.partial_ratio(texto_chunk, texto_macro) >= UMBRAL_FUZZY


def calcular_denominador(macro_id, source, chunking):
    """
    Cuenta cuántos chunks del documento `source` con la estrategia `chunking`
    están contenidos en la macro-sección dada.
    """
    indice = INDICES_POR_CHUNKING[chunking]
    texto_macro = mapa_macros[macro_id]

    # Recuperar TODOS los chunks del documento. Usamos scroll para no perder
    # nada si el documento tiene muchos chunks.
    chunks_doc = []
    resp = es.search(
        index=indice,
        body={
            'query': {'term': {'metadata.source': source}},
            '_source': ['text'],
        },
        size=1000,
        scroll='2m',
    )
    chunks_doc.extend(h['_source']['text'] for h in resp['hits']['hits'])
    scroll_id = resp.get('_scroll_id')
    while scroll_id and resp['hits']['hits']:
        resp = es.scroll(scroll_id=scroll_id, scroll='2m')
        chunks_doc.extend(h['_source']['text'] for h in resp['hits']['hits'])
    if scroll_id:
        try:
            es.clear_scroll(scroll_id=scroll_id)
        except Exception:
            pass

    if not chunks_doc:
        return 0

    return sum(1 for txt in chunks_doc if es_relevante(txt, texto_macro))


def precalcular_denominadores(resultados):
    """
    Calcula y cachea los denominadores de Recall.
    Clave: (macro_id, chunking). Independiente de la estrategia de retrieval.
    """
    if os.path.exists(ARCHIVO_CACHE_DENOMINADORES):
        with open(ARCHIVO_CACHE_DENOMINADORES, 'r', encoding='utf-8') as f:
            cache = json.load(f)
        print(f" - Denominadores cargados de caché: {len(cache)} entradas")
        return cache

    print(" - Precalculando denominadores de Recall (puede tardar)...")
    cache = {}
    pares_unicos = set()
    for r in resultados:
        if r['chunking'] not in CHUNKINGS_EVALUADOS or r['tipo'] not in TIPOS_EVALUADOS:
            continue
        pares_unicos.add((r['macro_id_origen'], r['source'], r['chunking']))

    for i, (macro_id, source, chunking) in enumerate(sorted(pares_unicos), 1):
        clave = f'{macro_id}__{chunking}'
        if clave in cache:
            continue
        n = calcular_denominador(macro_id, source, chunking)
        cache[clave] = n
        if i % 50 == 0:
            print(f"   {i}/{len(pares_unicos)} pares procesados...")

    with open(ARCHIVO_CACHE_DENOMINADORES, 'w', encoding='utf-8') as f:
        json.dump(cache, f, indent=2)
    print(f" - Denominadores guardados en caché: {ARCHIVO_CACHE_DENOMINADORES}")
    return cache


# ═══════════════════════════════════════════════════════════════════
# CÁLCULO DE MÉTRICAS POR EJECUCIÓN
# ═══════════════════════════════════════════════════════════════════

def evaluar_relevancias(chunks_recuperados, texto_macro):
    """Devuelve lista de booleanos: relevante o no por cada chunk."""
    return [es_relevante(ch['text'], texto_macro) for ch in chunks_recuperados]


def precision_at_k(relevancias, k):
    if k == 0:
        return 0.0
    rel_at_k = relevancias[:k]
    if not rel_at_k:
        return 0.0
    return sum(rel_at_k) / k


def recall_at_k(relevancias, k, total_relevantes):
    if total_relevantes == 0:
        return None  # no definible
    rel_at_k = relevancias[:k]
    return sum(rel_at_k) / total_relevantes


def mrr(relevancias):
    """Reciprocal rank del primer relevante."""
    for i, r in enumerate(relevancias, 1):
        if r:
            return 1.0 / i
    return 0.0


def ndcg_at_k(relevancias, k):
    """NDCG con relevancia binaria."""
    rel_at_k = relevancias[:k]
    if not rel_at_k:
        return 0.0

    # DCG: relevancia binaria, ganancia 1 / log2(rank+1)
    dcg = sum(rel / np.log2(i + 2) for i, rel in enumerate(rel_at_k))

    # IDCG: ranking ideal (todos los relevantes primero)
    n_rel = sum(relevancias)
    n_ideal = min(n_rel, k)
    if n_ideal == 0:
        return 0.0
    idcg = sum(1.0 / np.log2(i + 2) for i in range(n_ideal))
    return dcg / idcg


def calcular_metricas_ejecucion(r, denominadores_cache):
    """Calcula todas las métricas para una ejecución del pipeline."""
    macro_id = r['macro_id_origen']
    if macro_id not in mapa_macros:
        return None

    texto_macro = mapa_macros[macro_id]
    relevancias = evaluar_relevancias(r['chunks_recuperados'], texto_macro)

    clave_denom = f'{macro_id}__{r["chunking"]}'
    total_relevantes = denominadores_cache.get(clave_denom, 0)

    salida = {
        'pregunta_idx': r['pregunta_idx'],
        'tipo': r['tipo'],
        'source': r['source'],
        'chunking': r['chunking'],
        'retrieval': r['retrieval'],
        'macro_id_origen': macro_id,
        'n_relevantes_recuperados': sum(relevancias),
        'n_total_relevantes_corpus': total_relevantes,
        'mrr': mrr(relevancias),
    }

    for k in KS_EVALUACION:
        salida[f'P@{k}'] = precision_at_k(relevancias, k)
        salida[f'R@{k}'] = recall_at_k(relevancias, k, total_relevantes)
        salida[f'NDCG@{k}'] = ndcg_at_k(relevancias, k)

    return salida


# ═══════════════════════════════════════════════════════════════════
# AGREGACIÓN
# ═══════════════════════════════════════════════════════════════════

def agregar_metricas(df_metricas, agrupar_por):
    """Agrega métricas promediando sobre el grupo."""
    columnas_metricas = ['mrr']
    for k in KS_EVALUACION:
        columnas_metricas += [f'P@{k}', f'R@{k}', f'NDCG@{k}']

    return (df_metricas
            .groupby(agrupar_por)[columnas_metricas]
            .mean()
            .round(4)
            .reset_index())


# ═══════════════════════════════════════════════════════════════════
# EXCEL FORMATEADO
# ═══════════════════════════════════════════════════════════════════

def exportar_excel(df_global, df_por_tipo, df_por_chunking, df_por_retrieval):
    """Genera Excel con varias hojas y formato."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    thin = Side(style='thin', color='BBBBBB')
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    color_titulo = '1F3864'
    color_header = 'CCCCCC'

    def estilo(cell, bold=False, bg=None, color='000000'):
        cell.font = Font(name='Arial', bold=bold, color=color, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = borde
        if bg:
            cell.fill = PatternFill('solid', start_color=bg)

    def añadir_hoja(nombre, df, titulo):
        ws = wb.create_sheet(nombre)
        n_cols = len(df.columns)

        # Título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        c = ws.cell(row=1, column=1, value=titulo)
        estilo(c, bold=True, bg=color_titulo, color='FFFFFF')
        ws.row_dimensions[1].height = 24

        # Cabeceras
        for col, cab in enumerate(df.columns, 1):
            estilo(ws.cell(row=2, column=col, value=str(cab)),
                   bold=True, bg=color_header)
            ws.column_dimensions[get_column_letter(col)].width = 14

        # Datos
        for fila, (_, row) in enumerate(df.iterrows(), 3):
            for col, val in enumerate(row, 1):
                if isinstance(val, float):
                    val = round(val, 4) if pd.notna(val) else '-'
                estilo(ws.cell(row=fila, column=col, value=val))

    añadir_hoja('Global', df_global, 'MÉTRICAS IR — VISTA GLOBAL (chunking × retrieval)')
    añadir_hoja('Por Tipo', df_por_tipo, 'MÉTRICAS IR — DESGLOSE POR TIPO DE PREGUNTA')
    añadir_hoja('Por Chunking', df_por_chunking, 'MÉTRICAS IR — AGREGADO POR CHUNKING')
    añadir_hoja('Por Retrieval', df_por_retrieval, 'MÉTRICAS IR — AGREGADO POR RETRIEVAL')

    wb.save(ARCHIVO_SALIDA_EXCEL)
    print(f" - Excel guardado: {ARCHIVO_SALIDA_EXCEL}")


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════

def main():
    # Filtrar resultados a evaluar
    resultados_filtrados = [
        r for r in resultados
        if r['chunking'] in CHUNKINGS_EVALUADOS and r['tipo'] in TIPOS_EVALUADOS
    ]
    n_excluidas_chunking = sum(1 for r in resultados if r['chunking'] not in CHUNKINGS_EVALUADOS)
    n_excluidas_tipo = sum(1 for r in resultados
                            if r['chunking'] in CHUNKINGS_EVALUADOS and r['tipo'] not in TIPOS_EVALUADOS)

    print(f" - Ejecuciones totales:           {len(resultados)}")
    print(f" - Excluidas por chunking (D):    {n_excluidas_chunking}")
    print(f" - Excluidas por tipo (Negativa): {n_excluidas_tipo}")
    print(f" - A evaluar:                     {len(resultados_filtrados)}")
    print()

    # Precalcular denominadores
    denominadores = precalcular_denominadores(resultados_filtrados)
    print()

    # Calcular métricas por ejecución
    print(" - Calculando métricas por ejecución...")
    metricas_ejecuciones = []
    for r in resultados_filtrados:
        m = calcular_metricas_ejecucion(r, denominadores)
        if m is not None:
            metricas_ejecuciones.append(m)
    print(f"   Total ejecuciones evaluadas: {len(metricas_ejecuciones)}")

    # Guardar métricas detalladas
    with open(ARCHIVO_SALIDA_JSON, 'w', encoding='utf-8') as f:
        json.dump(metricas_ejecuciones, f, ensure_ascii=False, indent=2)
    print(f" - Métricas detalladas guardadas: {ARCHIVO_SALIDA_JSON}")

    # Agregaciones
    df = pd.DataFrame(metricas_ejecuciones)

    df_global = agregar_metricas(df, ['chunking', 'retrieval'])
    df_por_tipo = agregar_metricas(df, ['chunking', 'retrieval', 'tipo'])
    df_por_chunking = agregar_metricas(df, ['chunking'])
    df_por_retrieval = agregar_metricas(df, ['retrieval'])

    # Imprimir resumen global por consola
    print()
    print("═" * 70)
    print(" RESUMEN GLOBAL (chunking × retrieval)")
    print("═" * 70)
    print(df_global.to_string(index=False))

    print()
    print("═" * 70)
    print(" AGREGADO POR CHUNKING")
    print("═" * 70)
    print(df_por_chunking.to_string(index=False))

    print()
    print("═" * 70)
    print(" AGREGADO POR RETRIEVAL")
    print("═" * 70)
    print(df_por_retrieval.to_string(index=False))

    # Exportar Excel
    print()
    exportar_excel(df_global, df_por_tipo, df_por_chunking, df_por_retrieval)
    print()
    print("═" * 70)
    print(" Cálculo de métricas IR completado")
    print("═" * 70)


if __name__ == '__main__':
    main()
